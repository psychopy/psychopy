#!/usr/bin/env python
# -*- coding: utf-8 -*-

import weakref
import pickle
import os
import sys
import copy
import inspect
import codecs
import numpy as np
import pandas as pd
import json_tricks
from packaging.version import Version

import psychopy
from psychopy import logging
from psychopy import constants
from psychopy.tools.filetools import (openOutputFile, genDelimiter,
                                      genFilenameFromDelimiter, pathToString)
from psychopy.tools.fileerrortools import handleFileCollision
from psychopy.tools.arraytools import extendArr
from .utils import _getExcelCellName

try:
    import openpyxl
    if Version(openpyxl.__version__) >= Version('2.4.0'):
        # openpyxl moved get_column_letter to utils.cell
        from openpyxl.utils.cell import get_column_letter
    else:
        from openpyxl.cell import get_column_letter
    from openpyxl import load_workbook, Workbook
    haveOpenpyxl = True
except ImportError:
    haveOpenpyxl = False

_experiments = weakref.WeakValueDictionary()


class _ComparisonMixin():
    def __eq__(self, other):
        # NoneType and booleans, for example, don't have a .__dict__ attribute.
        try:
            getattr(other, '__dict__')
        except AttributeError:
            return False

        # Check if the dictionary keys are the same before proceeding.
        if set(self.__dict__.keys()) != set(other.__dict__.keys()):
            return False

        # Loop over all keys, implementing special handling for certain
        # data types.
        for key, val in self.__dict__.items():
            if isinstance(val, np.ma.core.MaskedArray):
                if not np.ma.allclose(val, getattr(other, key)):
                    return False
            elif isinstance(val, np.ndarray):
                if not np.allclose(val, getattr(other, key)):
                    return False
            elif isinstance(val, (pd.DataFrame, pd.Series)):
                if not val.equals(getattr(other, key)):
                    return False
            else:
                if val != getattr(other, key):
                    return False

        return True

    def __ne__(self, other):
        return not self == other


class _BaseTrialHandler(_ComparisonMixin):
    def setExp(self, exp):
        """Sets the ExperimentHandler that this handler is attached to

        Do NOT attempt to set the experiment using::

            trials._exp = myExperiment

        because it needs to be performed using the `weakref` module.
        """
        # need to use a weakref to avoid creating a circular reference that
        # prevents effective object deletion
        expId = id(exp)
        _experiments[expId] = exp
        self._exp = expId
        # origin will have been stored by the exp so don't store again:
        self.origin = None

    def getExp(self):
        """Return the ExperimentHandler that this handler is attached to,
        if any. Returns None if not attached
        """
        if self._exp is None or self._exp not in _experiments:
            return None
        else:
            return _experiments[self._exp]

    def _terminate(self):
        """Remove references to ourself in experiments and terminate the loop
        """
        # remove ourself from the list of unfinished loops in the experiment
        exp = self.getExp()
        if exp != None:
            exp.loopEnded(self)
        # and halt the loop
        raise StopIteration

    def saveAsPickle(self, fileName, fileCollisionMethod='rename'):
        """Basically just saves a copy of the handler (with data) to a
        pickle file.

        This can be reloaded if necessary and further analyses carried out.

        :Parameters:

            fileCollisionMethod: Collision method passed to
            :func:`~psychopy.tools.fileerrortools.handleFileCollision`
        """
        fileName = pathToString(fileName)

        if self.thisTrialN < 0 and self.thisRepN < 0:
            # if both are < 1 we haven't started
            if self.autoLog:
                logging.info('.saveAsPickle() called but no trials completed.'
                             ' Nothing saved')
            return -1

        if not fileName.endswith('.psydat'):
            fileName += '.psydat'

        with openOutputFile(fileName=fileName, append=False,
                            fileCollisionMethod=fileCollisionMethod) as f:
            pickle.dump(self, f)

        logging.info('saved data to %s' % f.name)

    def saveAsText(self, fileName,
                   stimOut=None,
                   dataOut=('n', 'all_mean', 'all_std', 'all_raw'),
                   delim=None,
                   matrixOnly=False,
                   appendFile=True,
                   summarised=True,
                   fileCollisionMethod='rename',
                   encoding='utf-8-sig'):
        """
        Write a text file with the data and various chosen stimulus attributes

        :Parameters:

        fileName:
            will have .tsv appended and can include path info.

        stimOut:
            the stimulus attributes to be output. To use this you need to
            use a list of dictionaries and give here the names of dictionary
            keys that you want as strings

        dataOut:
            a list of strings specifying the dataType and the analysis to
            be performed,in the form `dataType_analysis`. The data can be
            any of the types that you added using trialHandler.data.add()
            and the analysis can be either 'raw' or most things in the
            numpy library, including; 'mean','std','median','max','min'...
            The default values will output the raw, mean and std of all
            datatypes found

        delim:
            allows the user to use a delimiter other than tab
            ("," is popular with file extension ".csv")

        matrixOnly:
            outputs the data with no header row or extraInfo attached

        appendFile:
            will add this output to the end of the specified file if
            it already exists

        fileCollisionMethod:
            Collision method passed to
            :func:`~psychopy.tools.fileerrortools.handleFileCollision`

        encoding:
            The encoding to use when saving a the file. Defaults to `utf-8-sig`.

        """
        fileName = pathToString(fileName)

        if stimOut is None:
            stimOut = []

        if self.thisTrialN < 0 and self.thisRepN < 0:
            # if both are < 1 we haven't started
            if self.autoLog:
                logging.info('TrialHandler.saveAsText called but no trials'
                             ' completed. Nothing saved')
            return -1

        dataArray = self._createOutputArray(stimOut=stimOut,
                                            dataOut=dataOut,
                                            matrixOnly=matrixOnly)

        # set default delimiter if none given
        if delim is None:
            delim = genDelimiter(fileName)

        # create the file or send to stdout
        fileName = genFilenameFromDelimiter(fileName, delim)
        with openOutputFile(fileName=fileName, append=appendFile,
                            fileCollisionMethod=fileCollisionMethod,
                            encoding=encoding) as f:
            # loop through lines in the data matrix
            for line in dataArray:
                for cellN, entry in enumerate(line):
                    # surround in quotes to prevent effect of delimiter
                    if delim in str(entry):
                        f.write(u'"%s"' % str(entry))
                    else:
                        f.write(str(entry))
                    if cellN < (len(line) - 1):
                        f.write(delim)
                f.write("\n")  # add an EOL at end of each line

        if (fileName is not None) and (fileName != 'stdout') and self.autoLog:
            logging.info('saved data to %s' % f.name)

    def printAsText(self, stimOut=None,
                    dataOut=('all_mean', 'all_std', 'all_raw'),
                    delim='\t',
                    matrixOnly=False):
        """Exactly like saveAsText() except that the output goes
        to the screen instead of a file
        """
        if stimOut is None:
            stimOut = []
        self.saveAsText('stdout', stimOut, dataOut, delim, matrixOnly)

    def saveAsExcel(self, fileName, sheetName='rawData',
                    stimOut=None,
                    dataOut=('n', 'all_mean', 'all_std', 'all_raw'),
                    matrixOnly=False,
                    appendFile=True,
                    fileCollisionMethod='rename'):
        """
        Save a summary data file in Excel OpenXML format workbook
        (:term:`xlsx`) for processing in most spreadsheet packages.
        This format is compatible with versions of Excel (2007 or greater)
        and and with OpenOffice (>=3.0).

        It has the advantage over the simpler text files (see
        :func:`TrialHandler.saveAsText()` )
        that data can be stored in multiple named sheets within the file.
        So you could have a single file named after your experiment and
        then have one worksheet for each participant. Or you could have
        one file for each participant and then multiple sheets for
        repeated sessions etc.

        The file extension `.xlsx` will be added if not given already.

        :Parameters:

            fileName: string
                the name of the file to create or append. Can include
                relative or absolute path

            sheetName: string
                the name of the worksheet within the file

            stimOut: list of strings
                the attributes of the trial characteristics to be output.
                To use this you need to have provided a list of dictionaries
                specifying to trialList parameter of the TrialHandler and
                give here the names of strings specifying entries in that
                dictionary

            dataOut: list of strings
                specifying the dataType and the analysis to
                be performed, in the form `dataType_analysis`. The data
                can be any of the types that you added using
                trialHandler.data.add() and the analysis can be either
                'raw' or most things in the numpy library, including
                'mean','std','median','max','min'. e.g. `rt_max` will give
                a column of max reaction times across the trials assuming
                that `rt` values have been stored. The default values will
                output the raw, mean and std of all datatypes found.

            appendFile: True or False
                If False any existing file with this name will be
                kept and a new file will be created with a slightly different
                name. If you want to overwrite the old file, pass 'overwrite'
                to ``fileCollisionMethod``.
                If True then a new worksheet will be appended.
                If a worksheet already exists with that name a number will
                be added to make it unique.

            fileCollisionMethod: string
                Collision method (``rename``,``overwrite``, ``fail``) passed to
                :func:`~psychopy.tools.fileerrortools.handleFileCollision`
                This is ignored if ``append`` is ``True``.

        """
        fileName = pathToString(fileName)

        if stimOut is None:
            stimOut = []

        if self.thisTrialN < 0 and self.thisRepN < 0:
            # if both are < 1 we haven't started
            if self.autoLog:
                logging.info('TrialHandler.saveAsExcel called but no '
                             'trials completed. Nothing saved')
            return -1

        # NB this was based on the limited documentation (1 page wiki) for
        # openpyxl v1.0
        if not haveOpenpyxl:
            raise ImportError('openpyxl is required for saving files in'
                              ' Excel (xlsx) format, but was not found.')
            # return -1

        # create the data array to be sent to the Excel file
        dataArray = self._createOutputArray(stimOut=stimOut,
                                            dataOut=dataOut,
                                            matrixOnly=matrixOnly)

        if not fileName.endswith('.xlsx'):
            fileName += '.xlsx'
        # create or load the file
        if appendFile and os.path.isfile(fileName):
            wb = load_workbook(fileName)
            newWorkbook = False
        else:
            if not appendFile:
                # the file exists but we're not appending, a new file will
                # be saved with a slightly different name, unless 
                # fileCollisionMethod = ``overwrite``
                fileName = handleFileCollision(fileName,
                                               fileCollisionMethod)
            wb = Workbook()  # create new workbook
            wb.properties.creator = 'PsychoPy' + psychopy.__version__
            newWorkbook = True

        if newWorkbook:
            ws = wb.worksheets[0]
            ws.title = sheetName
        else:
            ws = wb.create_sheet()
            ws.title = sheetName

        # loop through lines in the data matrix
        for lineN, line in enumerate(dataArray):
            if line is None:
                continue
            for colN, entry in enumerate(line):
                if entry is None:
                    entry = ''
                try:
                    # if it can convert to a number (from numpy) then do it
                    val = float(entry)
                except Exception:
                    val = u"{}".format(entry)
                ws.cell(column=colN+1, row=lineN+1, value=val)

        wb.save(filename=fileName)

    def saveAsJson(self,
                   fileName=None,
                   encoding='utf-8-sig',
                   fileCollisionMethod='rename'):
        """
        Serialize the object to the JSON format.

        Parameters
        ----------
        fileName: string, or None
            the name of the file to create or append. Can include a relative or
            absolute path. If `None`, will not write to a file, but return an
            in-memory JSON object.

        encoding : string, optional
            The encoding to use when writing the file.

        fileCollisionMethod : string
            Collision method passed to
            :func:`~psychopy.tools.fileerrortools.handleFileCollision`. Can be
            either of `'rename'`, `'overwrite'`, or `'fail'`.

        Notes
        -----
        Currently, a copy of the object is created, and the copy's .origin
        attribute is set to an empty string before serializing
        because loading the created JSON file would sometimes fail otherwise.

        """
        fileName = pathToString(fileName)

        self_copy = copy.deepcopy(self)
        self_copy.origin = ''
        msg = ('Setting attribute .origin to empty string during JSON '
               'serialization.')
        logging.warn(msg)

        if (fileName is None) or (fileName == 'stdout'):
            return json_tricks.dumps(self_copy)
        else:
            with openOutputFile(fileName=fileName,
                                fileCollisionMethod=fileCollisionMethod,
                                encoding=encoding) as f:
                json_tricks.dump(self_copy, f)

            logging.info('Saved JSON data to %s' % f.name)

    def getOriginPathAndFile(self, originPath=None):
        """Attempts to determine the path of the script that created this
        data file and returns both the path to that script and its contents.
        Useful to store the entire experiment with the data.

        If originPath is provided (e.g. from Builder) then this is used
        otherwise the calling script is the originPath (fine from a
        standard python script).
        """
        # self.originPath and self.origin (the contents of the origin file)
        if originPath == -1:
            return -1, None  # the user wants to avoid storing this
        elif originPath is None or not os.path.isfile(originPath):
            try:
                originPath = inspect.getouterframes(
                    inspect.currentframe())[2][1]
                if self.autoLog:
                    logging.debug("Using %s as origin file" % originPath)
            except Exception:
                if self.autoLog:
                    logging.debug("Failed to find origin file using "
                                  "inspect.getouterframes")
                return '', ''
        if os.path.isfile(originPath):  # do we NOW have a path?
            with codecs.open(originPath, "r", encoding="utf-8-sig") as f:
                origin = f.read()
        else:
            origin = None
        return originPath, origin


class FileSaveMixin:
    """
    Mixin method for trial/experiment handlers which adds methods for handling file saving to 
    different formats.
    """

    # attributes for handling file collision
    _nextSaveCollision = {}
    appendFiles = False
    # attributes for column sorting
    sortColumns = False
    columnPriority = {}    

    def getAllEntries(self):
        """
        Should return a list of dicts corresponding to rows for data output
        """
        raise NotImplementedError()

    def getDataNames(self):
        """
        Should return an array of column headers for data output. Overload this to get more than 
        simply a combination of all the keys from getAllEntries.
        """
        df = pd.DataFrame(self.getAllEntries())

        return df.columns

    def getPriority(self, name):
        """
        Get the priority value for a given column. If no priority value is
        stored, returns best guess based on column name.

        Parameters
        ----------
        name : str
            Column name

        Returns
        -------
        int
            The priority value stored/guessed for this column, most likely a value from `constants.priority`, one of:
            - CRITICAL (30): Always at the start of the data file, generally reserved for Routine start times
            - HIGH (20): Important columns which are near the front of the data file
            - MEDIUM (10): Possibly important columns which are around the middle of the data file
            - LOW (0): Columns unlikely to be important which are at the end of the data file
            - EXCLUDE (-10): Always at the end of the data file, actively marked as unimportant
        """
        if name not in self.columnPriority:
            # store priority if not specified already
            self.columnPriority[name] = self._guessPriority(name)
        # return stored priority
        return self.columnPriority[name]

    def _guessPriority(self, name):
        """
        Get a best guess at the priority of a column based on its name

        Parameters
        ----------
        name : str
            Name of the column

        Returns
        -------
        int
            One of the following:
            - HIGH (19): Important columns which are near the front of the data file
            - MEDIUM (9): Possibly important columns which are around the middle of the data file
            - LOW (-1): Columns unlikely to be important which are at the end of the data file

            NOTE: Values returned from this function are 1 less than values in `constants.priority`,
            columns whose priority was guessed are behind equivalently prioritised columns whose priority
            was specified.
        """
        # if there's a dot, get attribute name
        if "." in name:
            name = name.split(".")[-1]

        # start off assuming low priority
        priority = constants.priority.LOW
        # if name is one of identified likely high priority columns, it's medium priority
        if name in [
            "keys", "rt", "x", "y", "leftButton", "numClicks", "numLooks", "clip", "response", "value",
            "frameRate", "participant"
        ]:
            priority = constants.priority.MEDIUM

        return priority - 1

    def setPriority(self, name, value=constants.priority.HIGH):
        """
        Set the priority of a column in the data file.

        Parameters
        ----------
        name : str
            Name of the column, e.g. `text.started`
        value : int
            Priority value to set the column to - higher priority columns appear nearer to the start of
            the data file. Use values from `constants.priority` as landmark values:
            - CRITICAL (30): Always at the start of the data file, generally reserved for Routine start times
            - HIGH (20): Important columns which are near the front of the data file
            - MEDIUM (10): Possibly important columns which are around the middle of the data file
            - LOW (0): Columns unlikely to be important which are at the end of the data file
            - EXCLUDE (-10): Always at the end of the data file, actively marked as unimportant
        """
        self.columnPriority[name] = value

    def queueNextCollision(self, fileCollisionMethod, fileName=None):
        """
        Tell this object that, next time the named file is saved, it should handle 
        collisions a certain way. This is useful if you want to save multiple times within an 
        experiment.

        Parameters
        ----------
        fileCollisionMethod : str
            File collision method to use, see `saveAsWideText` or `saveAsPickle` for 
            details.
        fileName : str
            Filename to queue collision on, if None (default) will use this ExperimentHandler's 
            `dataFileName`
        """
        # handle default
        if fileName is None:
            fileName = self.dataFileName
        # make filename iterable
        if not isinstance(fileName, (list, tuple)):
            fileName = [fileName]
        # queue collision
        for thisFileName in fileName:
            self._nextSaveCollision[thisFileName] = fileCollisionMethod

    def saveAsWideText(
        self,
        fileName,
        delim='auto',
        matrixOnly=False,
        appendFile=None,
        encoding='utf-8-sig',
        fileCollisionMethod=None,
        sortColumns=None
    ):
        """Saves a long, wide-format text file, with one line representing
        the attributes and data for a single trial. Suitable for analysis
        in R and SPSS.

        If `appendFile=True` then the data will be added to the bottom of
        an existing file. Otherwise, if the file exists already it will
        be kept and a new file will be created with a slightly different
        name. If you want to overwrite the old file, pass 'overwrite'
        to ``fileCollisionMethod``.

        If `matrixOnly=True` then the file will not contain a header row,
        which can be handy if you want to append data to an existing file
        of the same format.

        Parameters
        ----------

        fileName:
            if extension is not specified, '.csv' will be appended if
            the delimiter is ',', else '.tsv' will be appended.
            Can include path info.

        delim:
            allows the user to use a delimiter other than the default
            tab ("," is popular with file extension ".csv")

        matrixOnly:
            outputs the data with no header row.

        appendFile:
            will add this output to the end of the specified file if
            it already exists.

        encoding:
            The encoding to use when saving a the file.
            Defaults to `utf-8-sig`.

        fileCollisionMethod:
            Collision method passed to
            :func:`~psychopy.tools.fileerrortools.handleFileCollision`

        sortColumns : str or bool
            How (if at all) to sort columns in the data file. Can be:
            - "alphabetical", "alpha", "a" or True: Sort alphabetically by header name
            - "priority", "pr" or "p": Sort according to priority
            - other: Do not sort, columns remain in order they were added
        
        Returns
        -------
        str
            Final filename (including _1, _2, etc. and file extension) which data was saved as
        """
        # handle default delim
        delimOptions = {
            'comma': ",",
            'semicolon': ";",
            'tab': "\t"
        }
        if delim == 'auto' or delim is None:
            delim = genDelimiter(fileName)
        elif delim in delimOptions:
            delim = delimOptions[delim]
        # handle default sort option
        if sortColumns is None:
            sortColumns = self.sortColumns
        # handle default append mode
        if appendFile is None:
            appendFile = self.appendFiles
        # handle default file collision method
        if fileCollisionMethod is None and fileName in self._nextSaveCollision:
            fileCollisionMethod = self._nextSaveCollision.pop(fileName)
        elif fileCollisionMethod is None:
            fileCollisionMethod = "rename"
        # sanitize filename (make sure it has a delimiter)
        fileName = genFilenameFromDelimiter(fileName, delim)
        
        # create a pandas dataframe
        data = pd.DataFrame(
            self.getAllEntries(),
            columns=self.getDataNames()
        )
        # sort names as requested
        if sortColumns in ("alphabetical", "alpha", "a", True):
            # sort alphabetically
            data.columns = data.columns.sort()
        elif sortColumns in ("priority", "pr" or "p"):
            # map names to their priority
            priorityMap = []
            for name in data.columns:
                priority = self.columnPriority.get(name, self._guessPriority(name))
                priorityMap.append((priority, name))
            data.columns = [name for priority, name in sorted(priorityMap, reverse=True)]
        # open file
        f = openOutputFile(
            fileName, 
            append=appendFile,
            fileCollisionMethod=fileCollisionMethod,
            encoding=encoding
        )
        # write to file
        data.to_csv(
            f, 
            index=False, 
            sep=delim,
            header=(not matrixOnly)
        )
        # close file
        if f != sys.stdout:
            f.close()
        
        logging.info('Saved data to %r' % f.name)

        return fileName


class DataHandler(_ComparisonMixin, dict):
    """For handling data (used by TrialHandler, principally, rather than
    by users directly)

    Numeric data are stored as numpy masked arrays where the mask is set
    True for missing entries. When any non-numeric data (string, list or
    array) get inserted using DataHandler.add(val) the array is converted
    to a standard (not masked) numpy array with dtype='O' and where missing
    entries have value = "--".

    Attributes:
        - ['key']=data arrays containing values for that key
            (e.g. data['accuracy']=...)
        - dataShape=shape of data (x,y,...z,nReps)
        - dataTypes=list of keys as strings

    """
    def __init__(self, dataTypes=None, trials=None, dataShape=None):
        self.trials = trials
        self.dataTypes = []  # names will be added during addDataType
        self.isNumeric = {}
        # if given dataShape use it - otherwise guess!
        if dataShape:
            self.dataShape = dataShape
        elif self.trials:
            self.dataShape = list(np.asarray(trials.trialList, 'O').shape)
            self.dataShape.append(trials.nReps)

        # initialise arrays now if poss
        if dataTypes and self.dataShape:
            for thisType in dataTypes:
                self.addDataType(thisType)

    def __eq__(self, other):
        # We ignore an attached TrialHandler object, otherwise we will end up
        # in an infinite loop, as this DataHandler is attached to the
        # TrialHandler!

        from psychopy.data import TrialHandler

        if isinstance(self.trials, TrialHandler):
            self_copy = copy.deepcopy(self)
            other_copy = copy.deepcopy(other)
            del self_copy.trials, other_copy.trials
            result = super(DataHandler, self_copy).__eq__(other_copy)

            msg = ('TrialHandler object detected in .trials. Excluding it from '
                   'comparison.')
            logging.warning(msg)
        else:
            result = super(DataHandler, self).__eq__(other)

        return result

    def addDataType(self, names, shape=None):
        """Add a new key to the data dictionary of particular shape if
        specified (otherwise the shape of the trial matrix in the trial
        handler. Data are initialised to be zero everywhere. Not needed
        by user: appropriate types will be added during initialisation
        and as each xtra type is needed.
        """
        if not shape:
            shape = self.dataShape
        if not isinstance(names, str):
            # recursively call this function until we have a string
            for thisName in names:
                self.addDataType(thisName)
        else:
            # create the appropriate array in the dict
            # initially use numpy masked array of floats with mask=True
            # for missing vals. convert to a numpy array with dtype='O'
            # if non-numeric data given. NB don't use masked array with
            # dytpe='O' together - they don't unpickle
            self[names] = np.ma.zeros(shape, 'f')  # masked array of floats
            self[names].mask = True
            # add the name to the list
            self.dataTypes.append(names)
            self.isNumeric[names] = True  # until we need otherwise

    def add(self, thisType, value, position=None):
        """Add data to an existing data type (and add a new one if necess)
        """
        if not thisType in self:
            self.addDataType(thisType)
        if position is None:
            # 'ran' is always the first thing to update
            repN = sum(self['ran'][self.trials.thisIndex])
            if thisType != 'ran':
                # because it has already been updated
                repN -= 1
            # make a list where 1st digit is trial number
            position = [self.trials.thisIndex]
            position.append(repN)

        # check whether data falls within bounds
        posArr = np.asarray(position)
        shapeArr = np.asarray(self.dataShape)
        if not np.all(posArr < shapeArr):
            # array isn't big enough
            logging.warning('need a bigger array for: ' + thisType)
            # not implemented yet!
            self[thisType] = extendArr(self[thisType], posArr)
        # check for ndarrays with more than one value and for non-numeric data
        if (self.isNumeric[thisType] and
                ((type(value) == np.ndarray and len(value) > 1) or
                     (type(value) not in [float, int]))):
            self._convertToObjectArray(thisType)
        # insert the value
        self[thisType][position[0], int(position[1])] = value

    def _convertToObjectArray(self, thisType):
        """Convert this datatype from masked numeric array to unmasked
        object array
        """
        dat = self[thisType]
        # create an array of Object type
        self[thisType] = np.array(dat.data, dtype='O')
        # masked vals should be "--", others keep data
        # we have to repeat forcing to 'O' or text gets truncated to 4chars
        self[thisType] = np.where(dat.mask, '--', dat).astype('O')
        self.isNumeric[thisType] = False
