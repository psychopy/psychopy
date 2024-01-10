#!/usr/bin/env python
# -*- coding: utf-8 -*-

import sys
import os
import pickle
import copy
import warnings
import numpy as np
from pkg_resources import parse_version

import psychopy
from psychopy import logging
from psychopy.tools.filetools import openOutputFile, genDelimiter
from psychopy.tools.fileerrortools import handleFileCollision
from psychopy.contrib.quest import QuestObject
from psychopy.contrib.psi import PsiObject
from .base import _BaseTrialHandler, _ComparisonMixin
from .utils import _getExcelCellName

try:
    from collections.abc import Iterable
except ImportError:
    from collections import Iterable

try:
    # import openpyxl
    import openpyxl
    if parse_version(openpyxl.__version__) >= parse_version('2.4.0'):
        # openpyxl moved get_column_letter to utils.cell
        from openpyxl.utils.cell import get_column_letter
    else:
        from openpyxl.cell import get_column_letter
    from openpyxl.reader.excel import load_workbook
    haveOpenpyxl = True
except ImportError:
    haveOpenpyxl = False


class StairHandler(_BaseTrialHandler):
    """Class to handle smoothly the selection of the next trial
    and report current values etc.
    Calls to next() will fetch the next object given to this
    handler, according to the method specified.

    See ``Demos >> ExperimentalControl >> JND_staircase_exp.py``

    The staircase will terminate when *nTrials* AND *nReversals* have
    been exceeded. If *stepSizes* was an array and has been exceeded
    before nTrials is exceeded then the staircase will continue
    to reverse.

    *nUp* and *nDown* are always considered as 1 until the first reversal
    is reached. The values entered as arguments are then used.

    """

    def __init__(self,
                 startVal,
                 nReversals=None,
                 stepSizes=4,  # dB stepsize
                 nTrials=0,
                 nUp=1,
                 nDown=3,  # correct responses before stim goes down
                 applyInitialRule=True,
                 extraInfo=None,
                 method='2AFC',
                 stepType='db',
                 minVal=None,
                 maxVal=None,
                 originPath=None,
                 name='',
                 autoLog=True,
                 **kwargs):
        """
        :Parameters:

            startVal:
                The initial value for the staircase.

            nReversals:
                The minimum number of reversals permitted.
                If `stepSizes` is a list, but the minimum number of
                reversals to perform, `nReversals`, is less than the
                length of this list, PsychoPy will automatically increase
                the minimum number of reversals and emit a warning.

            stepSizes:
                The size of steps as a single value or a list (or array).
                For a single value the step size is fixed. For an array or
                list the step size will progress to the next entry
                at each reversal.

            nTrials:
                The minimum number of trials to be conducted. If the
                staircase has not reached the required number of reversals
                then it will continue.

            nUp:
                The number of 'incorrect' (or 0) responses before the
                staircase level increases.

            nDown:
                The number of 'correct' (or 1) responses before the
                staircase level decreases.

            applyInitialRule : bool
                Whether to apply a 1-up/1-down rule until the first reversal
                point (if `True`), before switching to the specified up/down
                rule.

            extraInfo:
                A dictionary (typically) that will be stored along with
                collected data using
                :func:`~psychopy.data.StairHandler.saveAsPickle` or
                :func:`~psychopy.data.StairHandler.saveAsText` methods.

            method:
                Not used and may be deprecated in future releases.

            stepType: *'db'*, 'lin', 'log'
                The type of steps that should be taken each time. 'lin'
                will simply add or subtract that amount each step, 'db'
                and 'log' will step by a certain number of decibels or
                log units (note that this will prevent your value ever
                reaching zero or less)

            minVal: *None*, or a number
                The smallest legal value for the staircase, which can be
                used to prevent it reaching impossible contrast values,
                for instance.

            maxVal: *None*, or a number
                The largest legal value for the staircase, which can be
                used to prevent it reaching impossible contrast values,
                for instance.

            Additional keyword arguments will be ignored.

        :Notes:

        The additional keyword arguments `**kwargs` might for example be
        passed by the `MultiStairHandler`, which expects a `label` keyword
        for each staircase. These parameters are to be ignored by the
        StairHandler.

        """
        self.name = name
        self.startVal = startVal
        self.nUp = nUp
        self.nDown = nDown
        self.applyInitialRule = applyInitialRule
        self.extraInfo = extraInfo
        self.method = method
        self.stepType = stepType

        try:
            self.stepSizes = list(stepSizes)
        except TypeError:
            # stepSizes is not array-like / iterable, i.e., a scalar.
            self.stepSizes = [stepSizes]

        self._variableStep = True if len(self.stepSizes) > 1 else False
        self.stepSizeCurrent = self.stepSizes[0]

        if nReversals is None:
            self.nReversals = len(self.stepSizes)
        elif len(self.stepSizes) > nReversals:
            msg = ('Increasing number of minimum required reversals to the '
                   'number of step sizes, (%i).' % len(self.stepSizes))
            logging.warn(msg)
            self.nReversals = len(self.stepSizes)
        else:
            self.nReversals = nReversals

        # to terminate the nTrials must be exceeded and either
        self.nTrials = nTrials
        self.finished = False
        self.thisTrialN = -1
        # a dict of lists where each should have the same length as the main
        # data:
        self.otherData = {}
        self.data = []
        self.intensities = []
        self.reversalPoints = []
        self.reversalIntensities = []
        # initially it goes down but on every step:
        self.currentDirection = 'start'
        # correct since last stim change (minus are incorrect):
        self.correctCounter = 0
        self.intensity = self.startVal
        self.minVal = minVal
        self.maxVal = maxVal
        self.autoLog = autoLog
        # a flag for the 1-up 1-down initial rule:
        self.initialRule = False

        # self.originPath and self.origin (the contents of the origin file)
        self.originPath, self.origin = self.getOriginPathAndFile(originPath)
        self._exp = None  # the experiment handler that owns me!

    def __iter__(self):
        return self

    @property
    def intensity(self):
        """The intensity (level) of the current staircase"""
        return self._nextIntensity

    @intensity.setter
    def intensity(self, intensity):
        """The intensity (level) of the current staircase"""
        self._nextIntensity = intensity

    def addResponse(self, result, intensity=None):
        """Add a 1 or 0 to signify a correct / detected or
        incorrect / missed trial.

        This is essential to advance the staircase to a new intensity level!

        Supplying an `intensity` value here indicates that you did not use
        the recommended intensity in your last trial and the staircase will
        replace its recorded value with the one you supplied here.
        """
        self.data.append(result)

        # if needed replace the existing intensity with this custom one
        if intensity != None:
            self.intensities.pop()
            self.intensities.append(intensity)

        # increment the counter of correct scores
        if result == 1:
            if len(self.data) > 1 and self.data[-2] == result:
                # increment if on a run
                self.correctCounter += 1
            else:
                # or reset
                self.correctCounter = 1
        else:
            if len(self.data) > 1 and self.data[-2] == result:
                # increment if on a run
                self.correctCounter -= 1
            else:
                # or reset
                self.correctCounter = -1

        # add the current data to experiment if poss
        if self.getExp() is not None:  # update the experiment handler too
            self.getExp().addData(self.name + ".response", result)
        self.calculateNextIntensity()

    def addOtherData(self, dataName, value):
        """Add additional data to the handler, to be tracked alongside
        the result data but not affecting the value of the staircase
        """
        if not dataName in self.otherData:  # init the list
            if self.thisTrialN > 0:
                # might have run trials already
                self.otherData[dataName] = [None] * (self.thisTrialN - 1)
            else:
                self.otherData[dataName] = []
        # then add current value
        self.otherData[dataName].append(value)
        # add the current data to experiment if poss
        if self.getExp() != None:  # update the experiment handler too
            self.getExp().addData(dataName, value)

    def addData(self, result, intensity=None):
        """Deprecated since 1.79.00: This function name was ambiguous.
        Please use one of these instead:

        *   .addResponse(result, intensity)
        *   .addOtherData('dataName', value')

        """
        self.addResponse(result, intensity)

    def calculateNextIntensity(self):
        """Based on current intensity, counter of correct responses, and
        current direction.
        """

        if not self.reversalIntensities and self.applyInitialRule:
            # always using a 1-down, 1-up rule initially
            if self.data[-1] == 1:  # last answer correct
                # got it right
                if self.currentDirection == 'up':
                    reversal = True
                else:
                    # direction is 'down' or 'start'
                    reversal = False
                self.currentDirection = 'down'
            else:
                # got it wrong
                if self.currentDirection == 'down':
                    reversal = True
                else:
                    # direction is 'up' or 'start'
                    reversal = False
                # now:
                self.currentDirection = 'up'
        elif self.correctCounter >= self.nDown:
            # n right, time to go down!
            # 'start' covers `applyInitialRule=False`.
            if self.currentDirection not in ['start', 'down']:
                reversal = True
            else:
                reversal = False
            self.currentDirection = 'down'
        elif self.correctCounter <= -self.nUp:
            # n wrong, time to go up!
            # note current direction
            # 'start' covers `applyInitialRule=False`.
            if self.currentDirection not in ['start', 'up']:
                reversal = True
            else:
                reversal = False
            self.currentDirection = 'up'
        else:
            # same as previous trial
            reversal = False

        # add reversal info
        if reversal:
            self.reversalPoints.append(self.thisTrialN)
            if not self.reversalIntensities and self.applyInitialRule:
                self.initialRule = True
            self.reversalIntensities.append(self.intensities[-1])

        # test if we're done
        if (len(self.reversalIntensities) >= self.nReversals and
                    len(self.intensities) >= self.nTrials):
            self.finished = True

        # new step size if necessary
        if reversal and self._variableStep:
            if len(self.reversalIntensities) >= len(self.stepSizes):
                # we've gone beyond the list of step sizes
                # so just use the last one
                self.stepSizeCurrent = self.stepSizes[-1]
            else:
                _sz = len(self.reversalIntensities)
                self.stepSizeCurrent = self.stepSizes[_sz]

        # apply new step size
        if ((not self.reversalIntensities or self.initialRule) and
                self.applyInitialRule):
            self.initialRule = False  # reset the flag
            if self.data[-1] == 1:
                self._intensityDec()
            else:
                self._intensityInc()
        elif self.correctCounter >= self.nDown:
            # n right, so going down
            self._intensityDec()
        elif self.correctCounter <= -self.nUp:
            # n wrong, so going up
            self._intensityInc()

    def __next__(self):
        """Advances to next trial and returns it.
        Updates attributes; `thisTrial`, `thisTrialN` and `thisIndex`.

        If the trials have ended, calling this method will raise a
        StopIteration error. This can be handled with code such as::

            staircase = data.StairHandler(.......)
            for eachTrial in staircase:  # automatically stops when done
                # do stuff

        or::

            staircase = data.StairHandler(.......)
            while True:  # ie forever
                try:
                    thisTrial = staircase.next()
                except StopIteration:  # we got a StopIteration error
                    break  # break out of the forever loop
                # do stuff here for the trial

        """
        if not self.finished:
            # check that all 'otherData' is aligned with current trialN
            for key in self.otherData:
                while len(self.otherData[key]) < self.thisTrialN:
                    self.otherData[key].append(None)
            # update pointer for next trial
            self.thisTrialN += 1
            self.intensities.append(self._nextIntensity)
            return self._nextIntensity
        else:
            self._terminate()

    next = __next__  # allows user to call without a loop `val = trials.next()`

    def _intensityInc(self):
        """increment the current intensity and reset counter
        """
        if self.stepType == 'db':
            self._nextIntensity *= 10.0**(self.stepSizeCurrent/20.0)
        elif self.stepType == 'log':
            self._nextIntensity *= 10.0**self.stepSizeCurrent
        elif self.stepType == 'lin':
            self._nextIntensity += self.stepSizeCurrent
        # check we haven't gone out of the legal range
        if (self.maxVal is not None) and (self._nextIntensity > self.maxVal):
            self._nextIntensity = self.maxVal
        self.correctCounter = 0

    def _intensityDec(self):
        """decrement the current intensity and reset counter
        """
        if self.stepType == 'db':
            self._nextIntensity /= 10.0**(self.stepSizeCurrent/20.0)
        if self.stepType == 'log':
            self._nextIntensity /= 10.0**self.stepSizeCurrent
        elif self.stepType == 'lin':
            self._nextIntensity -= self.stepSizeCurrent
        self.correctCounter = 0
        # check we haven't gone out of the legal range
        if (self.minVal is not None) and (self._nextIntensity < self.minVal):
            self._nextIntensity = self.minVal

    def saveAsText(self, fileName,
                   delim=None,
                   matrixOnly=False,
                   fileCollisionMethod='rename',
                   encoding='utf-8-sig'):
        """Write a text file with the data

        :Parameters:

            fileName: a string
                The name of the file, including path if needed. The extension
                `.tsv` will be added if not included.

            delim: a string
                the delimitter to be used (e.g. '\t' for tab-delimitted,
                ',' for csv files)

            matrixOnly: True/False
                If True, prevents the output of the `extraInfo` provided
                at initialisation.

            fileCollisionMethod:
                Collision method passed to
                :func:`~psychopy.tools.fileerrortools.handleFileCollision`

            encoding:
                The encoding to use when saving a the file.
                Defaults to `utf-8-sig`.

        """

        if self.thisTrialN < 1:
            if self.autoLog:
                logging.debug('StairHandler.saveAsText called but no '
                              'trials completed. Nothing saved')
            return -1

        # set default delimiter if none given
        if delim is None:
            delim = genDelimiter(fileName)

        # create the file or send to stdout
        f = openOutputFile(
            fileName, append=False, 
            fileCollisionMethod=fileCollisionMethod, encoding=encoding)

        # write the data
        reversalStr = str(self.reversalIntensities)
        reversalStr = reversalStr.replace(',', delim)
        reversalStr = reversalStr.replace('[', '')
        reversalStr = reversalStr.replace(']', '')
        f.write('\nreversalIntensities=\t%s\n' % reversalStr)

        reversalPts = str(self.reversalPoints)
        reversalPts = reversalPts.replace(',', delim)
        reversalPts = reversalPts.replace('[', '')
        reversalPts = reversalPts.replace(']', '')
        f.write('reversalIndices=\t%s\n' % reversalPts)

        rawIntens = str(self.intensities)
        rawIntens = rawIntens.replace(',', delim)
        rawIntens = rawIntens.replace('[', '')
        rawIntens = rawIntens.replace(']', '')
        f.write('\nintensities=\t%s\n' % rawIntens)

        responses = str(self.data)
        responses = responses.replace(',', delim)
        responses = responses.replace('[', '')
        responses = responses.replace(']', '')
        f.write('responses=\t%s\n' % responses)

        # add self.extraInfo
        if self.extraInfo is not None and not matrixOnly:
            strInfo = str(self.extraInfo)
            # dict begins and ends with {} - remove
            # strInfo.replace('{','')
            # strInfo = strInfo.replace('}','')
            strInfo = strInfo[1:-1]
            # separate value from keyname
            strInfo = strInfo.replace(strInfo, ': ', ':\n')
            # separate values from each other
            strInfo = strInfo.replace(',', '\n')
            strInfo = strInfo.replace('array([ ', '')
            strInfo = strInfo.replace('])', '')

            f.write('\n%s\n' % strInfo)

        f.write("\n")
        if f != sys.stdout:
            f.close()
            if self.autoLog:
                logging.info('saved data to %s' % f.name)

    def saveAsExcel(self, fileName, sheetName='data',
                    matrixOnly=False, appendFile=True,
                    fileCollisionMethod='rename'):
        """Save a summary data file in Excel OpenXML format workbook
        (:term:`xlsx`) for processing in most spreadsheet packages.
        This format is compatible with versions of Excel (2007 or greater)
        and and with OpenOffice (>=3.0).

        It has the advantage over the simpler text files
        (see :func:`TrialHandler.saveAsText()` ) that data can be stored
        in multiple named sheets within the file. So you could have a
        single file named after your experiment and then have one worksheet
        for each participant. Or you could have one file for each participant
        and then multiple sheets for repeated sessions etc.

        The file extension `.xlsx` will be added if not given already.

        The file will contain a set of values specifying the staircase level
        ('intensity') at each reversal, a list of reversal indices
        (trial numbers), the raw staircase / intensity level on *every*
        trial and the corresponding responses of the participant on every
        trial.

        :Parameters:

            fileName: string
                the name of the file to create or append. Can include
                relative or absolute path.

            sheetName: string
                the name of the worksheet within the file

            matrixOnly: True or False
                If set to True then only the data itself will be output
                (no additional info)

            appendFile: True or False
                If False any existing file with this name will be
                overwritten. If True then a new worksheet will be appended.
                If a worksheet already exists with that name a number will
                be added to make it unique.

            fileCollisionMethod: string
                Collision method passed to
                :func:`~psychopy.tools.fileerrortools.handleFileCollision`
                This is ignored if ``appendFile`` is ``True``.

        """

        if self.thisTrialN < 1:
            if self.autoLog:
                logging.debug('StairHandler.saveAsExcel called but no '
                              'trials completed. Nothing saved')
            return -1
        # NB this was based on the limited documentation for openpyxl v1.0
        if not haveOpenpyxl:
            raise ImportError('openpyxl is required for saving files in '
                              'Excel (xlsx) format, but was not found.')
            # return -1

        # import necessary subpackages - they are small so won't matter to do
        # it here
        from openpyxl.workbook import Workbook
        from openpyxl.reader.excel import load_workbook

        if not fileName.endswith('.xlsx'):
            fileName += '.xlsx'
        # create or load the file
        if appendFile and os.path.isfile(fileName):
            wb = load_workbook(fileName)
            newWorkbook = False
        else:
            if not appendFile:
                # the file exists but we're not appending, will be overwritten
                fileName = handleFileCollision(fileName,
                                               fileCollisionMethod)
            wb = Workbook()
            wb.properties.creator = 'PsychoPy' + psychopy.__version__
            newWorkbook = True

        if newWorkbook:
            ws = wb.worksheets[0]
            ws.title = sheetName
        else:
            ws = wb.create_sheet()
            ws.title = sheetName

        # write the data
        # reversals data
        ws['A1'] = 'Reversal Intensities'
        ws['B1'] = 'Reversal Indices'
        for revN, revIntens in enumerate(self.reversalIntensities):
            ws.cell(column=1, row=revN+2,
                    value=u"{}".format(revIntens))
            ws.cell(column=2, row=revN+2,
                    value=u"{}".format(self.reversalPoints[revN]))

        # trials data
        ws['C1'] = 'All Intensities'
        ws['D1'] = 'All Responses'
        for intenN, intensity in enumerate(self.intensities):
            ws.cell(column=3, row=intenN+2,
                    value=u"{}".format(intensity))
            ws.cell(column=4, row=intenN+2,
                    value=u"{}".format(self.data[intenN]))

        # add other data
        col = 5
        if self.otherData is not None:
            # for varName in self.otherData:
            for key, val in list(self.otherData.items()):
                ws.cell(column=col, row=1,
                        value=u"{}".format(key))
                for oDatN in range(len(self.otherData[key])):
                    ws.cell(column=col, row=oDatN+2,
                            value=u"{}".format(self.otherData[key][oDatN]))
                col += 1

        # add self.extraInfo
        if self.extraInfo is not None and not matrixOnly:
            ws.cell(column=col, row=1,
                    value='extraInfo')
            rowN = 2
            for key, val in list(self.extraInfo.items()):
                ws.cell(column=col, row=rowN,
                        value=u"{}:".format(key))
                _cell = _getExcelCellName(col=col+1, row=rowN)
                ws.cell(column=col+2, row=rowN+1,
                        value=u"{}".format(val))
                rowN += 1


        wb.save(filename=fileName)
        if self.autoLog:
            logging.info('saved data to %s' % fileName)

    def saveAsPickle(self, fileName, fileCollisionMethod='rename'):
        """Basically just saves a copy of self (with data) to a pickle file.

        This can be reloaded if necessary and further analyses carried out.

        :Parameters:

            fileCollisionMethod: Collision method passed to
            :func:`~psychopy.tools.fileerrortools.handleFileCollision`

        """
        if self.thisTrialN < 1:
            if self.autoLog:
                logging.debug('StairHandler.saveAsPickle called but no '
                              'trials completed. Nothing saved')
            return -1

        # otherwise use default location
        if not fileName.endswith('.psydat'):
            fileName += '.psydat'

        with openOutputFile(fileName=fileName, append=False,
                            fileCollisionMethod=fileCollisionMethod) as f:
            pickle.dump(self, f)

        if (fileName is not None) and (fileName != 'stdout'):
            logging.info('saved data to %s' % f.name)


class QuestObject_(QuestObject, _ComparisonMixin):
    """A QuestObject that implements the == and != operators.
    """
    pass


class QuestHandler(StairHandler):
    r"""Class that implements the Quest algorithm for quick measurement of
    psychophysical thresholds.

    Uses Andrew Straw's `QUEST <http://www.visionegg.org/Quest>`_, which is a
    Python port of Denis Pelli's Matlab code.

    Measures threshold using a Weibull psychometric function. Currently, it is
    not possible to use a different psychometric function.

    The Weibull psychometric function is given by the formula
    
    :math:`\Psi(x) = \delta \gamma + (1 - \delta) [1 - (1 - \gamma)\, \exp(-10^{\\beta (x - T + \epsilon)})]`

    Here, :math:`x` is an intensity or a contrast (in log10 units), and :math:`T` is estimated threshold.
    
    Quest internally shifts the psychometric function such that intensity at the user-specified
    threshold performance level ``pThreshold`` (e.g., 50% in a yes-no or 75% in a 2-AFC task) is euqal to 0.
    The parameter :math:`\epsilon` is responsible for this shift, and is determined automatically based on the
    specified ``pThreshold`` value. It is the parameter Watson & Pelli (1983) introduced to perform measurements
    at the "optimal sweat factor". Assuming your ``QuestHandler`` instance is called ``q``, you can retrieve this
    value via ``q.epsilon``.

    **Example**::

        # setup display/window
        ...
        # create stimulus
        stimulus = visual.RadialStim(win=win, tex='sinXsin', size=1,
                                     pos=[0,0], units='deg')
        ...
        # create staircase object
        # trying to find out the point where subject's response is 50 / 50
        # if wanted to do a 2AFC then the defaults for pThreshold and gamma
        # are good. As start value, we'll use 50% contrast, with SD = 20%
        staircase = data.QuestHandler(0.5, 0.2,
            pThreshold=0.63, gamma=0.01,
            nTrials=20, minVal=0, maxVal=1)
        ...
        while thisContrast in staircase:
            # setup stimulus
            stimulus.setContrast(thisContrast)
            stimulus.draw()
            win.flip()
            core.wait(0.5)
            # get response
            ...
            # inform QUEST of the response, needed to calculate next level
            staircase.addResponse(thisResp)
        ...
        # can now access 1 of 3 suggested threshold levels
        staircase.mean()
        staircase.mode()
        staircase.quantile(0.5)  # gets the median

    """

    def __init__(self,
                 startVal,
                 startValSd,
                 pThreshold=0.82,
                 nTrials=None,
                 stopInterval=None,
                 method='quantile',
                 beta=3.5,
                 delta=0.01,
                 gamma=0.5,
                 grain=0.01,
                 range=None,
                 extraInfo=None,
                 minVal=None,
                 maxVal=None,
                 staircase=None,
                 originPath=None,
                 name='',
                 autoLog=True,
                 **kwargs):
        """
        Typical values for pThreshold are:
            * 0.82 which is equivalent to a 3 up 1 down standard staircase
            * 0.63 which is equivalent to a 1 up 1 down standard staircase
                (and might want gamma=0.01)

        The variable(s) nTrials and/or stopSd must be specified.

        `beta`, `delta`, and `gamma` are the parameters of the Weibull
        psychometric function.

        :Parameters:

            startVal:
                Prior threshold estimate or your initial guess threshold.

            startValSd:
                Standard deviation of your starting guess threshold.
                Be generous with the sd as QUEST will have trouble finding
                the true threshold if it's more than one sd from your
                initial guess.

            pThreshold
                Your threshold criterion expressed as probability of
                response==1. An intensity offset is introduced into the
                psychometric function so that the threshold (i.e.,
                the midpoint of the table) yields pThreshold.

            nTrials: *None* or a number
                The maximum number of trials to be conducted.

            stopInterval: *None* or a number
                The minimum 5-95% confidence interval required in the
                threshold estimate before stopping. If both this and
                nTrials is specified, whichever happens first will
                determine when Quest will stop.

            method: *'quantile'*, 'mean', 'mode'
                The method used to determine the next threshold to test.
                If you want to get a specific threshold level at the end
                of your staircasing, please use the quantile, mean, and
                mode methods directly.

            beta: *3.5* or a number
                Controls the steepness of the psychometric function.

            delta: *0.01* or a number
                The fraction of trials on which the observer presses blindly.

            gamma: *0.5* or a number
                The fraction of trials that will generate response 1 when
                intensity=-Inf.

            grain: *0.01* or a number
                The quantization of the internal table.

            range: *None*, or a number
                The intensity difference between the largest and smallest
                intensity that the internal table can store. This interval
                will be centered on the initial guess tGuess. QUEST assumes
                that intensities outside of this range have zero prior
                probability (i.e., they are impossible).

            extraInfo:
                A dictionary (typically) that will be stored along with
                collected data using
                :func:`~psychopy.data.StairHandler.saveAsPickle` or
                :func:`~psychopy.data.StairHandler.saveAsText` methods.

            minVal: *None*, or a number
                The smallest legal value for the staircase, which can be
                used to prevent it reaching impossible contrast values,
                for instance.

            maxVal: *None*, or a number
                The largest legal value for the staircase, which can be
                used to prevent it reaching impossible contrast values,
                for instance.

            staircase: *None* or StairHandler
                Can supply a staircase object with intensities and results.
                Might be useful to give the quest algorithm more information
                if you have it. You can also call the importData function
                directly.

            Additional keyword arguments will be ignored.

        :Notes:

        The additional keyword arguments `**kwargs` might for example be
        passed by the `MultiStairHandler`, which expects a `label` keyword
        for each staircase. These parameters are to be ignored by the
        StairHandler.

        """
        StairHandler.__init__(
            self, startVal, nTrials=nTrials, extraInfo=extraInfo,
            method=method, stepType='lin', minVal=minVal,
            maxVal=maxVal, name=name, autoLog=autoLog)

        self.startVal = startVal
        self.startValSd = startValSd
        self.pThreshold = pThreshold
        self.stopInterval = stopInterval
        # NB there is also _nextIntensity
        self._questNextIntensity = startVal
        self._range = range

        # Create Quest object
        self._quest = QuestObject_(
            startVal, startValSd, pThreshold, beta, delta, gamma,
            grain=grain, range=self._range)

        # Import any old staircase data
        if staircase is not None:
            self.importData(staircase.intensities, staircase.data)
        # store the origin file and its path
        self.originPath, self.origin = self.getOriginPathAndFile(originPath)
        self._exp = None
        self.autoLog = autoLog

    # NB we inherit self.intensity from StairHandler

    @property
    def beta(self):
        return self._quest.beta

    @property
    def gamma(self):
        return self._quest.gamma

    @property
    def delta(self):
        return self._quest.delta
    @property
    def epsilon(self):
        return self._quest.xThreshold

    @property
    def grain(self):
        return self._quest.grain

    @property
    def range(self):
        return self._range

    def addResponse(self, result, intensity=None):
        """Add a 1 or 0 to signify a correct / detected or
        incorrect / missed trial

        Supplying an `intensity` value here indicates that you did not use the
        recommended intensity in your last trial and the staircase will
        replace its recorded value with the one you supplied here.
        """
        # Process user supplied intensity
        if intensity is None:
            intensity = self._questNextIntensity
        else:
            # Update the intensity.
            #
            # During the first trial, self.intensities will be of length 0,
            # so pop() would not work.
            if len(self.intensities) != 0:
                self.intensities.pop()  # remove the auto-generated one
            self.intensities.append(intensity)
        # Update quest
        self._quest.update(intensity, result)
        # Update other things
        self.data.append(result)
        # add the current data to experiment if poss
        if self.getExp() != None:  # update the experiment handler too
            self.getExp().addData(self.name + ".response", result)

        self._checkFinished()
        if not self.finished:
            self.calculateNextIntensity()

    def importData(self, intensities, results):
        """import some data which wasn't previously given to the quest
        algorithm
        """
        # NOT SURE ABOUT CLASS TO USE FOR RAISING ERROR
        if len(intensities) != len(results):
            raise AttributeError("length of intensities and results input "
                                 "must be the same")
        self.incTrials(len(intensities))
        for intensity, result in zip(intensities, results):
            try:
                next(self)
                self.addResponse(result, intensity)
            except StopIteration:
                # would get a stop iteration if stopInterval set
                pass    # TODO: might want to check if nTrials is still good

    def calculateNextIntensity(self):
        """based on current intensity and counter of correct responses
        """
        self._intensity()
        # Check we haven't gone out of the legal range
        if self.maxVal is not None and self._nextIntensity > self.maxVal:
            self._nextIntensity = self.maxVal
        elif self.minVal is not None and self._nextIntensity < self.minVal:
            self._nextIntensity = self.minVal
        self._questNextIntensity = self._nextIntensity

    def _intensity(self):
        """assigns the next intensity level"""
        if self.method == 'mean':
            self._questNextIntensity = self._quest.mean()
        elif self.method == 'mode':
            self._questNextIntensity = self._quest.mode()[0]
        elif self.method == 'quantile':
            self._questNextIntensity = self._quest.quantile()
        else:
            raise TypeError(f"Requested method for QUEST: {self.method} is not a valid method. Please use mean, mode or quantile")
        self._nextIntensity = self._questNextIntensity

    def mean(self):
        """mean of Quest posterior pdf
        """
        return self._quest.mean()

    def sd(self):
        """standard deviation of Quest posterior pdf
        """
        return self._quest.sd()

    def mode(self):
        """mode of Quest posterior pdf
        """
        return self._quest.mode()[0]

    def quantile(self, p=None):
        """quantile of Quest posterior pdf
        """
        return self._quest.quantile(quantileOrder=p)

    def confInterval(self, getDifference=False):
        """
        Return estimate for the 5%--95% confidence interval (CI).

        :Parameters:

            getDifference (bool)
                If ``True``, return the width of the confidence interval
                (95% - 5% percentiles). If ``False``, return an NumPy array
                with estimates for the 5% and 95% boundaries.

        :Returns:

            scalar or array of length 2.
        """
        interval = [self.quantile(0.05), self.quantile(0.95)]
        if getDifference:
            return abs(interval[0] - interval[1])
        else:
            return interval

    def incTrials(self, nNewTrials):
        """increase maximum number of trials
        Updates attribute: `nTrials`
        """
        self.nTrials += nNewTrials

    def simulate(self, tActual):
        """returns a simulated user response to the next intensity level
        presented by Quest, need to supply the actual threshold level
        """
        # Current estimated intensity level
        if self.method == 'mean':
            tTest = self._quest.mean()
        elif self.method == 'mode':
            tTest = self._quest.mode()
        elif self.method == 'quantile':
            tTest = self._quest.quantile()
        return self._quest.simulate(tTest, tActual)

    def __next__(self):
        """Advances to next trial and returns it.
        Updates attributes; `thisTrial`, `thisTrialN`, `thisIndex`,
        `finished`, `intensities`

        If the trials have ended, calling this method will raise a
        StopIteration error. This can be handled with code such as::

            staircase = data.QuestHandler(.......)
            for eachTrial in staircase:  # automatically stops when done
                # do stuff

        or::

            staircase = data.QuestHandler(.......)
            while True:  # i.e. forever
                try:
                    thisTrial = staircase.next()
                except StopIteration:  # we got a StopIteration error
                    break  # break out of the forever loop
                # do stuff here for the trial
        """
        if self.finished == False:
            # update pointer for next trial
            self.thisTrialN += 1
            self.intensities.append(self._nextIntensity)
            return self._nextIntensity
        else:
            self._terminate()

    next = __next__  # allows user to call without a loop `val = trials.next()`

    def _checkFinished(self):
        """checks if we are finished
        Updates attribute: `finished`
        """
        if self.nTrials is not None and len(self.intensities) >= self.nTrials:
            self.finished = True
        elif (self.stopInterval is not None and
                      self.confInterval(True) < self.stopInterval):
            self.finished = True
        else:
            self.finished = False


class PsiObject_(PsiObject, _ComparisonMixin):
    """A PsiObject that implements the == and != operators.
    """
    pass


class PsiHandler(StairHandler):
    """Handler to implement the "Psi" adaptive psychophysical method
    (Kontsevich & Tyler, 1999).

    This implementation assumes the form of the psychometric function
    to be a cumulative Gaussian. Psi estimates the two free parameters
    of the psychometric function, the location (alpha) and slope (beta),
    using Bayes' rule and grid approximation of the posterior distribution.
    It chooses stimuli to present by minimizing the entropy of this grid.
    Because this grid is represented internally as a 4-D array, one must
    choose the intensity, alpha, and beta ranges carefully so as to avoid
    a Memory Error. Maximum likelihood is used to estimate Lambda, the most
    likely location/slope pair. Because Psi estimates the entire
    psychometric function, any threshold defined on the function may be
    estimated once Lambda is determined.

    It is advised that Lambda estimates are examined after completion of
    the Psi procedure. If the estimated alpha or beta values equal your
    specified search bounds, then the search range most likely did not
    contain the true value. In this situation the procedure should be
    repeated with appropriately adjusted bounds.

    Because Psi is a Bayesian method, it can be initialized with a prior
    from existing research. A function to save the posterior over Lambda
    as a Numpy binary file is included.

    Kontsevich & Tyler (1999) specify their psychometric function in terms
    of d'. PsiHandler avoids this and treats all parameters with respect
    to stimulus intensity. Specifically, the forms of the psychometric
    function assumed for Yes/No and Two Alternative Forced Choice (2AFC)
    are, respectively:

    _normCdf = norm.cdf(x, mean=alpha, sd=beta)
    Y(x) = .5 * delta + (1 - delta) * _normCdf

    Y(x) = .5 * delta + (1 - delta) * (.5 + .5 * _normCdf)
    """

    def __init__(self,
                 nTrials,
                 intensRange, alphaRange, betaRange,
                 intensPrecision, alphaPrecision, betaPrecision,
                 delta,
                 stepType='lin',
                 expectedMin=0.5,
                 prior=None,
                 fromFile=False,
                 extraInfo=None,
                 name=''):
        """Initializes the handler and creates an internal Psi Object for
        grid approximation.

        :Parameters:

            nTrials (int)
                The number of trials to run.

            intensRange (list)
                Two element list containing the (inclusive) endpoints of
                the stimuli intensity range.

            alphaRange  (list)
                Two element list containing the (inclusive) endpoints of
                the alpha (location parameter) range.

            betaRange   (list)
                Two element list containing the (inclusive) endpoints of
                the beta (slope parameter) range.

            intensPrecision (float or int)
                If stepType == 'lin', this specifies the step size of the
                stimuli intensity range. If stepType == 'log', this specifies
                the number of steps in the stimuli intensity range.

            alphaPrecision  (float)
                The step size of the alpha (location parameter) range.

            betaPrecision   (float)
                The step size of the beta (slope parameter) range.

            delta   (float)
                The guess rate.

            stepType    (str)
                The type of steps to be used when constructing the stimuli
                intensity range. If 'lin' then evenly spaced steps are used.
                If 'log' then logarithmically spaced steps are used.
                Defaults to 'lin'.

            expectedMin  (float)
                The expected lower asymptote of the psychometric function
                (PMF).

                For a Yes/No task, the PMF usually extends across the
                interval [0, 1]; here, `expectedMin` should be set to `0`.

                For a 2-AFC task, the PMF spreads out across [0.5, 1.0].
                Therefore, `expectedMin` should be set to `0.5` in this
                case, and the 2-AFC psychometric function described above
                going to be is used.

                Currently, only Yes/No and 2-AFC designs are supported.

                Defaults to 0.5, or a 2-AFC task.

            prior   (numpy ndarray or str)
                Optional prior distribution with which to initialize the
                Psi Object. This can either be a numpy ndarray object or
                the path to a numpy binary file (.npy) containing the ndarray.

            fromFile    (str)
                Flag specifying whether prior is a file pathname or not.

            extraInfo   (dict)
                Optional dictionary object used in PsychoPy's built-in
                logging system.

            name    (str)
                Optional name for the PsiHandler used in PsychoPy's built-in
                logging system.

        :Raises:

            NotImplementedError
                If the supplied `minVal` parameter implies an experimental
                design other than Yes/No or 2-AFC.

        """
        if expectedMin not in [0, 0.5]:
            raise NotImplementedError(
                'Currently, only Yes/No and 2-AFC designs are '
                'supported. Please specify either `expectedMin=0` '
                '(Yes/No) or `expectedMin=0.5` (2-AFC).')

        StairHandler.__init__(
            self, startVal=None, nTrials=nTrials, extraInfo=extraInfo,
            stepType=stepType, minVal=intensRange[0],
            maxVal=intensRange[1], name=name
        )

        # Create Psi object
        if prior is not None and fromFile:
            try:
                prior = np.load(prior)
            except IOError:
                logging.warning("The specified pickle file could not be "
                                "read. Using a uniform prior instead.")
                prior = None

        twoAFC = True if expectedMin == 0.5 else False
        self._psi = PsiObject_(
            intensRange, alphaRange, betaRange, intensPrecision,
            alphaPrecision, betaPrecision, delta=delta,
            stepType=stepType, TwoAFC=twoAFC, prior=prior)

        self._psi.update(None)

    def addResponse(self, result, intensity=None):
        """Add a 1 or 0 to signify a correct / detected or
        incorrect / missed trial. Supplying an `intensity` value here
        indicates that you did not use the
        recommended intensity in your last trial and the staircase will
        replace its recorded value with the one you supplied here.
        """
        self.data.append(result)

        # if needed replace the existing intensity with this custom one
        if intensity is not None:
            self.intensities.pop()
            self.intensities.append(intensity)
        # add the current data to experiment if possible
        if self.getExp() is not None:
            # update the experiment handler too
            self.getExp().addData(self.name + ".response", result)
        self._psi.update(result)

    def __next__(self):
        """Advances to next trial and returns it.
        """
        self._checkFinished()
        if self.finished == False:
            # update pointer for next trial
            self.thisTrialN += 1
            self.intensities.append(self._psi.nextIntensity)
            return self._psi.nextIntensity
        else:
            self._terminate()

    next = __next__  # allows user to call without a loop `val = trials.next()`

    def _checkFinished(self):
        """checks if we are finished.
        Updates attribute: `finished`
        """
        if self.nTrials is not None and len(self.intensities) >= self.nTrials:
            self.finished = True
        else:
            self.finished = False

    def estimateLambda(self):
        """Returns a tuple of (location, slope)
        """
        return self._psi.estimateLambda()

    def estimateThreshold(self, thresh, lamb=None):
        """Returns an intensity estimate for the provided probability.

        The optional argument 'lamb' allows thresholds to be estimated
        without having to recompute the maximum likelihood lambda.
        """
        if lamb is not None:
            try:
                if len(lamb) != 2:
                    msg = ("Invalid user-specified lambda pair. A "
                           "new estimate of lambda will be computed.")
                    warnings.warn(msg, SyntaxWarning)
                    lamb = None
            except TypeError:
                msg = ("Invalid user-specified lambda pair. A new "
                       "estimate of lambda will be computed.")
                warnings.warn(msg, SyntaxWarning)
                lamb = None
        return self._psi.estimateThreshold(thresh, lamb)

    def savePosterior(self, fileName, fileCollisionMethod='rename'):
        """Saves the posterior array over probLambda as a pickle file
        with the specified name.

        Parameters
        ----------
        fileCollisionMethod : string
            Collision method passed to :func:`~psychopy.tools.fileerrortools.handleFileCollision`

        """
        try:
            if os.path.exists(fileName):
                fileName = handleFileCollision(
                    fileName,
                    fileCollisionMethod=fileCollisionMethod
                )
            self._psi.savePosterior(fileName)
        except IOError:
            warnings.warn("An error occurred while trying to save the "
                          "posterior array. Continuing without saving...")


class QuestPlusHandler(StairHandler):
    def __init__(self,
                 nTrials,
                 intensityVals, thresholdVals, slopeVals,
                 lowerAsymptoteVals, lapseRateVals,
                 responseVals=('Yes', 'No'), prior=None,
                 startIntensity=None,
                 psychometricFunc='weibull', stimScale='log10',
                 stimSelectionMethod='minEntropy',
                 stimSelectionOptions=None, paramEstimationMethod='mean',
                 extraInfo=None, name='', label='', **kwargs):
        """
        QUEST+ implementation. Currently only supports parameter estimation of
        a Weibull-shaped psychometric function.

        The parameter estimates can be retrieved via the `.paramEstimate`
        attribute, which returns a dictionary whose keys correspond to the
        names of the estimated parameters (i.e., `QuestPlusHandler.paramEstimate['threshold']`
        will provide the threshold estimate). Retrieval of the marginal posterior distributions works
        similarly: they can be accessed via the `.posterior` dictionary.

        Parameters
        ----------
        nTrials : int
            Number of trials to run.

        intensityVals : collection of floats
            The complete set of possible stimulus levels. Note that the
            stimulus levels are not necessarily limited to intensities (as the
            name of this parameter implies), but they could also be contrasts,
            durations, weights, etc.

        thresholdVals : float or collection of floats
            The complete set of possible threshold values.

        slopeVals : float or collection of floats
            The complete set of possible slope values.

        lowerAsymptoteVals : float or collection of floats
            The complete set of possible values of the lower asymptote. This
            corresponds to false-alarm rates in yes-no tasks, and to the
            guessing rate in n-AFC tasks. Therefore, when performing an n-AFC
            experiment, the collection should consists of a single value only
            (e.g., `[0.5]` for 2-AFC, `[0.33]` for 3-AFC, `[0.25]` for 4-AFC,
            etc.).

        lapseRateVals : float or collection of floats
            The complete set of possible lapse rate values. The lapse rate
            defines the upper asymptote of the psychometric function, which
            will be at `1 - lapse rate`.

        responseVals : collection
            The complete set of possible response outcomes. Currently, only
            two outcomes are supported: the first element must correspond to
            a successful response / stimulus detection, and the second one to
            an unsuccessful or incorrect response. For example, in a yes-no
            task, one would use `['Yes', 'No']`, and in an n-AFC task,
            `['Correct', 'Incorrect']`; or, alternatively, the less verbose
            `[1, 0]` in both cases.

        prior : dict of floats
            The prior probabilities to assign to the parameter values. The
            dictionary keys correspond to the respective parameters:
            ``threshold``, ``slope``, ``lowerAsymptote``, ``lapseRate``.

        startIntensity : float
            The very first intensity (or stimulus level) to present.

        psychometricFunc : {'weibull'}
            The psychometric function to fit. Currently, only the Weibull
            function is supported.

        stimScale : {'log10', 'dB', 'linear'}
            The scale on which the stimulus intensities (or stimulus levels)
            are provided. Currently supported are the decadic logarithm,
            `log10`; decibels, `dB`; and a linear scale, `linear`.

        stimSelectionMethod : {'minEntropy', 'minNEntropy'}
            How to select the next stimulus. `minEntropy` will select the
            stimulus that will minimize the expected entropy. `minNEntropy`
            will randomly pick pick a stimulus from the set of stimuli that
            will produce the smallest, 2nd-smallest, ..., N-smallest entropy.
            This can be used to ensure some variation in the stimulus selection
            (and subsequent presentation) procedure. The number `N` will then
            have to be specified via the `stimSelectionOption` parameter.

        stimSelectionOptions : dict
            This parameter further controls how to select the next stimulus in
            case `stimSelectionMethod=minNEntropy`.
            The dictionary supports two keys:
            `N` and `maxConsecutiveReps`.
            `N` defines the number of "best" stimuli (i.e., those which
            produce the smallest `N` expected entropies) from which to randomly
            select a stimulus for presentation in the next trial.
            `maxConsecutiveReps` defines how many times the exact same stimulus
            can be presented on consecutive trials.
            For example, to randomly pick a stimulus from those which will
            produce the 4 smallest expected entropies, and to allow the same
            stimulus to be presented on two consecutive trials max, use
            `stimSelectionOptions=dict(N=4, maxConsecutiveReps=2)`.
            To achieve reproducible results, you may pass a seed to the
            random number generator via the `randomSeed` key.

        paramEstimationMethod : {'mean', 'mode'}
            How to calculate the final parameter estimate. `mean` returns the
            mean of each parameter, weighted by their respective posterior
            probabilities. `mode` returns the parameters at the peak of
            the posterior distribution.

        extraInfo : dict
            Additional information to store along the actual QUEST+ staircase
            data.

        name : str
            The name of the QUEST+ staircase object. This will appear in the
            PsychoPy logs.

        label : str
            Only used by :class:`MultiStairHandler`, and otherwise ignored.

        kwargs : dict
            Additional keyword arguments. These might be passed, for example,
            through a :class:`MultiStairHandler`, and will be ignored. A
            warning will be emitted whenever additional keyword arguments
            have been passed.

        Warns
        -----
        RuntimeWarning
            If an unknown keyword argument was passed.

        Notes
        -----
        The QUEST+ algorithm was first described by [1]_.

        .. [1] Andrew B. Watson (2017). QUEST+: A general multidimensional
               Bayesian adaptive psychometric method.
               Journal of Vision, 17(3):10. doi: 10.1167/17.3.10.

        """
        if sys.version_info.major == 3 and sys.version_info.minor >= 6:
            import questplus as qp
        else:
            msg = 'QUEST+ implementation requires Python 3.6 or newer'
            raise RuntimeError(msg)

        msg = ('The QUEST+ staircase implementation is currently being '
               'tested and may be subject to change.')
        logging.critical(msg)

        if kwargs:
            msg = ('The  following keyword arguments are unknown to '
                   'QuestPlusHandler and will be ignored: \n')
            for k in kwargs.keys():
                msg += '\n  - %s' % k
            msg += ('\n\nIf you are using QuestPlusHandler through a '
                    'MultiStairHandler, it may be safe to ignore this '
                    'warning.')
            logging.warn(msg)

            # Ensure we get a proper unit-testable warning too (not just a
            # logfile entry)
            msg = 'Unknown keyword argument(s) passed to QuestPlusHandler'
            warnings.warn(msg, RuntimeWarning)

        super().__init__(startVal=startIntensity, nTrials=nTrials,
                         stepType=stimScale, extraInfo=extraInfo, name=name)

        # We  don't use these attributes that were inherited from StairHandler.
        self.currentDirection = None
        self.stepSizeCurrent = None
        # Note that self.stepType is not used either: we use self.stimScale
        # instead (which is defined below).

        self.intensityVals = intensityVals
        self.thresholdVals = thresholdVals
        self.slopeVals = slopeVals
        self.lowerAsymptoteVals = lowerAsymptoteVals
        self.lapseRateVals = lapseRateVals
        self.responseVals = responseVals

        self.psychometricFunc = psychometricFunc
        self.stimScale = stimScale
        self.stimSelectionMethod = stimSelectionMethod
        self.stimSelectionOptions = stimSelectionOptions
        self.paramEstimationMethod = paramEstimationMethod
        self._prior = prior

        # questplus uses different parameter names.
        if self.stimSelectionMethod == 'minEntropy':
            stimSelectionMethod_ = 'min_entropy'
        elif self.stimSelectionMethod == 'minNEntropy':
            stimSelectionMethod_ = 'min_n_entropy'
        else:
            raise ValueError('Unknown stimSelectionMethod requested.')

        if self.stimSelectionOptions is not None:
            valid = ('N', 'maxConsecutiveReps', 'randomSeed')
            if any([o not in valid for o in self.stimSelectionOptions]):
                msg = ('Unknown stimSelectionOptions requested. '
                       'Valid options are: %s' % ', '.join(valid))
                raise ValueError(msg)

            stimSelectionOptions_ = dict()

            if 'N' in self.stimSelectionOptions:
                stimSelectionOptions_['n'] = self.stimSelectionOptions['N']
            if 'maxConsecutiveReps' in self.stimSelectionOptions:
                stimSelectionOptions_['max_consecutive_reps'] = self.stimSelectionOptions['maxConsecutiveReps']
            if 'randomSeed' in self.stimSelectionOptions:
                stimSelectionOptions_['random_seed'] = self.stimSelectionOptions['randomSeed']
        else:
            stimSelectionOptions_ = self.stimSelectionOptions

        if self._prior is not None:
            valid = ('threshold', 'slope', 'lapseRate', 'lowerAsymptote')
            if any([p not in valid for p in self._prior]):
                msg = ('Invalid prior parameter(s) specified. '
                       'Valid parameter names are: %s' % ', '.join(valid))
                raise ValueError(msg)

            prior_ = dict()

            if 'threshold' in self._prior:
                prior_['threshold'] = self._prior['threshold']
            if 'slope' in self._prior:
                prior_['slope'] = self._prior['slope']
            if 'lapseRate' in self._prior:
                prior_['lapse_rate'] = self._prior['lapseRate']
            if 'lowerAsymptote' in self._prior:
                prior_['lower_asymptote'] = self._prior['lowerAsymptote']
        else:
            prior_ = self._prior

        if self.psychometricFunc == 'weibull':
            self._qp = qp.QuestPlusWeibull(
                intensities=self.intensityVals,
                thresholds=self.thresholdVals,
                slopes=self.slopeVals,
                lower_asymptotes=self.lowerAsymptoteVals,
                lapse_rates=self.lapseRateVals,
                prior=prior_,
                responses=self.responseVals,
                stim_scale=self.stimScale,
                stim_selection_method=stimSelectionMethod_,
                stim_selection_options=stimSelectionOptions_,
                param_estimation_method=self.paramEstimationMethod)
        else:
            msg = ('Currently only the Weibull psychometric function is '
                   'supported.')
            raise ValueError(msg)

        # Ensure self._nextIntensity is set in case the `startIntensity` kwarg
        # was supplied. We never actually use self._nextIntensity in the
        # QuestPlusHandler; it's mere purpose here is to make the
        # MultiStairHandler happy.
        if self.startIntensity is not None:
            self._nextIntensity = self.startIntensity
        else:
            self._nextIntensity = self._qp.next_intensity

    @property
    def startIntensity(self):
        return self.startVal

    def addResponse(self, response, intensity=None):
        self.data.append(response)

        # if needed replace the existing intensity with this custom one
        if intensity is not None:
            self.intensities.pop()
            self.intensities.append(intensity)
        # add the current data to experiment if possible
        if self.getExp() is not None:
            # update the experiment handler too
            self.getExp().addData(self.name + ".response", response)
        self._qp.update(intensity=self.intensities[-1],
                        response=response)

    def __next__(self):
        self._checkFinished()
        if not self.finished:
            # update pointer for next trial
            self.thisTrialN += 1
            if self.thisTrialN == 0 and self.startIntensity is not None:
                self.intensities.append(self.startVal)
            else:
                self.intensities.append(self._qp.next_intensity)

            # We never actually use self._nextIntensity in the
            # QuestPlusHandler; it's mere purpose here is to make the
            # MultiStairHandler happy.
            self._nextIntensity = self.intensities[-1]
            return self.intensities[-1]
        else:
            self._terminate()

    next = __next__

    def _checkFinished(self):
        if self.nTrials is not None and len(self.intensities) >= self.nTrials:
            self.finished = True
        else:
            self.finished = False

    @property
    def paramEstimate(self):
        """
        The estimated parameters of the psychometric function.

        Returns
        -------
        dict of floats
            A dictionary whose keys correspond to the names of the estimated
            parameters.

        """
        qp_estimate = self._qp.param_estimate
        estimate = dict(threshold=qp_estimate['threshold'],
                        slope=qp_estimate['slope'],
                        lowerAsymptote=qp_estimate['lower_asymptote'],
                        lapseRate=qp_estimate['lapse_rate'])
        return estimate

    @property
    def prior(self):
        """
        The marginal prior distributions.

        Returns
        -------
        dict of np.ndarrays
            A dictionary whose keys correspond to the names of the parameters.

        """
        qp_prior = self._qp.prior

        threshold = qp_prior.sum(dim=('slope', 'lower_asymptote', 'lapse_rate'))
        slope = qp_prior.sum(dim=('threshold', 'lower_asymptote', 'lapse_rate'))
        lowerAsymptote = qp_prior.sum(dim=('threshold', 'slope', 'lapse_rate'))
        lapseRate = qp_prior.sum(dim=('threshold', 'slope', 'lower_asymptote'))

        qp_prior = dict(threshold=threshold.values,
                        slope=slope.values,
                        lowerAsymptote=lowerAsymptote.values,
                        lapseRate=lapseRate.values)
        return qp_prior

    @property
    def posterior(self):
        """
        The marginal posterior distributions.

        Returns
        -------
        dict of np.ndarrays
            A dictionary whose keys correspond to the names of the estimated
            parameters.

        """
        qp_posterior = self._qp.posterior

        threshold = qp_posterior.sum(dim=('slope', 'lower_asymptote', 'lapse_rate'))
        slope = qp_posterior.sum(dim=('threshold', 'lower_asymptote', 'lapse_rate'))
        lowerAsymptote = qp_posterior.sum(dim=('threshold', 'slope', 'lapse_rate'))
        lapseRate = qp_posterior.sum(dim=('threshold', 'slope', 'lower_asymptote'))

        posterior = dict(threshold=threshold.values,
                         slope=slope.values,
                         lowerAsymptote=lowerAsymptote.values,
                         lapseRate=lapseRate.values)
        return posterior

    def saveAsJson(self,
                   fileName=None,
                   encoding='utf-8-sig',
                   fileCollisionMethod='rename'):
        self_copy = copy.deepcopy(self)

        # Convert questplus.QuestPlus to JSON using questplus's built-in
        # functionality. questplus uses xarray, which cannot be easily
        # serialized directly using json_tricks (yet).
        self_copy._qp_json = self_copy._qp.to_json()
        del self_copy._qp

        r = (super(QuestPlusHandler, self_copy)
             .saveAsJson(fileName=fileName,
                         encoding=encoding,
                         fileCollisionMethod=fileCollisionMethod))

        if fileName is None:
            return r


class MultiStairHandler(_BaseTrialHandler):

    def __init__(self, stairType='simple', method='random',
                 conditions=None, nTrials=50, randomSeed=None,
                 originPath=None, name='', autoLog=True):
        """A Handler to allow easy interleaved staircase procedures
        (simple or QUEST).

        Parameters for the staircases, as used by the relevant
        :class:`StairHandler` or
        :class:`QuestHandler` (e.g. the `startVal`, `minVal`, `maxVal`...)
        should be specified in the `conditions` list and may vary between
        each staircase. In particular, the conditions **must** include
        a `startVal` (because this is a required argument to the above
        handlers), a `label` to tag the staircase and a `startValSd`
        (only for QUEST staircases). Any parameters not specified in the
        conditions file will revert to the default for that individual
        handler.

        If you need to customize the behaviour further you may want to
        look at the recipe on :ref:`interleavedStairs`.

        :params:

            stairType: 'simple', 'quest', or 'questplus'
                Use a :class:`StairHandler`, a :class:`QuestHandler`, or a
                 :class:`QuestPlusHandler`.

            method: 'random', 'fullRandom', or 'sequential'
                If `random`, stairs are shuffled in each repeat but not
                randomized more than that (so you can't have 3 repeats of the
                same staircase in a row unless it's the only one still
                running). If `fullRandom`, the staircase order is "fully"
                randomized, meaning that, theoretically, a large number of
                subsequent trials could invoke the same staircase repeatedly.
                If `sequential`, don't perform any randomization.

            conditions: a list of dictionaries specifying conditions
                Can be used to control parameters for the different staircases.
                Can be imported from an Excel file using
                `psychopy.data.importConditions`
                MUST include keys providing, 'startVal', 'label' and
                'startValSd' (QUEST only).
                The 'label' will be used in data file saving so should
                be unique.
                See Example Usage below.

            nTrials=50
                Minimum trials to run (but may take more if the staircase
                hasn't also met its minimal reversals.
                See :class:`~psychopy.data.StairHandler`

            randomSeed : int or None
                The seed with which to initialize the random number generator
                (RNG). If `None` (default), do not initialize the RNG with
                a specific value.

        Example usage::

            conditions=[
                {'label':'low', 'startVal': 0.1, 'ori':45},
                {'label':'high','startVal': 0.8, 'ori':45},
                {'label':'low', 'startVal': 0.1, 'ori':90},
                {'label':'high','startVal': 0.8, 'ori':90},
                ]
            stairs = data.MultiStairHandler(conditions=conditions, nTrials=50)

            for thisIntensity, thisCondition in stairs:
                thisOri = thisCondition['ori']

                # do something with thisIntensity and thisOri

                stairs.addResponse(correctIncorrect)  # this is ESSENTIAL

            # save data as multiple formats
            stairs.saveDataAsExcel(fileName)  # easy to browse
            stairs.saveAsPickle(fileName)  # contains more info

        Raises
        ------
            ValueError
                If an unknown randomization option was passed via the `method`
                keyword argument.

        """
        self.name = name
        self.autoLog = autoLog
        self.type = stairType
        self.method = method
        self.randomSeed = randomSeed
        self._rng = np.random.RandomState(seed=randomSeed)
        self.conditions = conditions
        self.nTrials = nTrials
        self.finished = False
        self.totalTrials = 0
        self._checkArguments()
        # create staircases
        self.staircases = []  # all staircases
        self.runningStaircases = []  # staircases that haven't finished yet
        self.thisPassRemaining = []  # staircases to run this pass
        self._createStairs()

        # fetch first staircase/value (without altering/advancing it)
        self._startNewPass()
        self.currentStaircase = self.thisPassRemaining[0]  # take the first
        # gets updated by self.addData()
        self._nextIntensity = self.currentStaircase._nextIntensity
        # store the origin file and its path
        self.originPath, self.origin = self.getOriginPathAndFile(originPath)
        self._exp = None  # the experiment handler that owns me!

    def _checkArguments(self):
        # Did we get a `conditions` parameter, correctly formatted?
        if not isinstance(self.conditions, Iterable):
            raise TypeError(
                '`conditions` parameter passed to MultiStairHandler '
                'should be a list, not a %s.' % type(self.conditions))

        c0 = self.conditions[0]
        if not isinstance(c0, dict):
            raise TypeError(
                '`conditions` passed to MultiStairHandler should be a '
                'list of python dictionaries, not a list of %ss.' %
                type(c0))

        # Did `conditions` contain the things we need?
        params = list(c0.keys())
        if self.type not in ['simple', 'quest', 'QUEST', 'questplus']:
            raise ValueError(
                'MultiStairHandler `stairType` should be \'simple\', '
                '\'QUEST\' or \'quest\', not \'%s\'' % self.type)

        if self.type != 'questplus' and 'startVal' not in params:
            raise AttributeError('MultiStairHandler needs a parameter called '
                                 '`startVal` in conditions')
        if 'label' not in params:
            raise AttributeError('MultiStairHandler needs a parameter called'
                                 ' `label` in conditions')
        if self.type in ['QUEST', 'quest'] and 'startValSd' not in params:
            raise AttributeError(
                'MultiStairHandler needs a parameter called '
                '`startValSd` in conditions for QUEST staircases.')

    def _createStairs(self):
        for condition in self.conditions:
            # We create a copy, because we are going to remove items from
            # this dictionary in this loop, but don't want these
            # changes to alter the originals in self.conditions.
            args = dict(condition)

            # If no individual `nTrials` parameter was supplied for this
            # staircase, use the `nTrials` that were passed to
            # the MultiStairHandler on instantiation.
            if 'nTrials' not in args:
                args['nTrials'] = self.nTrials

            if self.type == 'simple':
                startVal = args.pop('startVal')
                thisStair = StairHandler(startVal, **args)
            elif self.type in ['QUEST', 'quest']:
                startVal = args.pop('startVal')
                startValSd = args.pop('startValSd')
                thisStair = QuestHandler(startVal, startValSd, **args)
            elif self.type == 'questplus':
                thisStair = QuestPlusHandler(**args)

            # This isn't normally part of handler.
            thisStair.condition = condition

            # And finally, add it to the list.
            self.staircases.append(thisStair)
            self.runningStaircases.append(thisStair)

    def __iter__(self):
        return self

    def __next__(self):
        """Advances to next trial and returns it.

        This can be handled with code such as::

            staircase = data.MultiStairHandler(.......)
            for eachTrial in staircase:  # automatically stops when done
                # do stuff here for the trial

        or::

            staircase = data.MultiStairHandler(.......)
            while True:  # ie forever
                try:
                    thisTrial = staircase.next()
                except StopIteration:  # we got a StopIteration error
                    break  # break out of the forever loop
                # do stuff here for the trial

        """
        # create a new set for this pass if needed
        if (not hasattr(self, 'thisPassRemaining') or
                not self.thisPassRemaining):
            if self.runningStaircases:
                self._startNewPass()
            else:
                self.finished = True
                raise StopIteration

        # fetch next staircase/value
        self.currentStaircase = self.thisPassRemaining.pop(
            0)  # take the first and remove it
        # if staircase.next() not called, staircaseHandler would not
        # save the first intensity,
        # Error: miss align intensities and responses
        # gets updated by self.addResponse()
        self._nextIntensity = next(self.currentStaircase)

        # return value
        if not self.finished:
            # inform experiment of the condition (but not intensity,
            # that might be overridden by user)
            if self.getExp() != None:
                exp = self.getExp()
                stair = self.currentStaircase
                for key, value in list(stair.condition.items()):
                    exp.addData("%s.%s" % (self.name, key), value)
                exp.addData(self.name + '.thisIndex',
                            self.conditions.index(stair.condition))
                exp.addData(self.name + '.thisRepN', stair.thisTrialN + 1)
                exp.addData(self.name + '.thisN', self.totalTrials)

                if self.type != 'questplus':
                    exp.addData(self.name + '.direction', stair.currentDirection)
                    exp.addData(self.name + '.stepSize', stair.stepSizeCurrent)
                    exp.addData(self.name + '.stepType', stair.stepType)

                exp.addData(self.name + '.intensity', self._nextIntensity)

            self._trialAborted = False  # reset this flag
            return self._nextIntensity, self.currentStaircase.condition
        else:
            raise StopIteration

    next = __next__  # allows user to call without a loop `val = trials.next()`

    def _startNewPass(self):
        """Create a new iteration of the running staircases for this pass.

        This is not normally needed by the user - it gets called at __init__
        and every time that next() runs out of trials for this pass.
        """
        if self.method == 'sequential':
            self.thisPassRemaining = copy.copy(self.runningStaircases)
        elif self.method == 'random':
            # np.random.shuffle() works in-place!
            self.thisPassRemaining = copy.copy(self.runningStaircases)
            self._rng.shuffle(self.thisPassRemaining)
        elif self.method == 'fullRandom':
            n = len(self.runningStaircases)
            self.thisPassRemaining = self._rng.choice(self.runningStaircases,
                                                      size=n, replace=True)
            # np.random.choice() returns an ndarray, so convert back to a list
            # again.
            self.thisPassRemaining = list(self.thisPassRemaining)
        else:
            raise ValueError('Unknown randomization method requested.')

    @property
    def intensity(self):
        """The intensity (level) of the current staircase"""
        return self.currentStaircase._nextIntensity

    @intensity.setter
    def intensity(self, intensity):
        """The intensity (level) of the current staircase"""
        self.currentStaircase._nextIntensity = intensity

    def abortCurrentTrial(self, action='random'):
        """Abort the current trial (staircase).

        Calling this during an experiment abort the current staircase used this
        trial. The current staircase will be reshuffled into available 
        staircases depending on the `action` parameter.

        Parameters
        ----------
        action : str
            Action to take with the aborted trial. Can be either of `'random'`,
            or `'append'`. The default action is `'random'`.

        Notes
        -----
        * When using `action='random'`, the RNG state for the trial handler is
          not used.

        """
        # check if value for parameter `action` is valid
        if not isinstance(action, str):  # type checks for params
            raise TypeError(
                "Parameter `action` specified incorrect type, must be `str`.")
        
        # reinsert the current staircase into the list of running staircases
        if action == 'append':
            self.thisPassRemaining.append(self.currentStaircase)
        elif action == 'random':
            self.thisPassRemaining.append(self.currentStaircase)
            # shuffle using the numpy RNG to preserve state
            np.random.shuffle(self.thisPassRemaining)
        else:
            raise ValueError(
                "Value for parameter `action` must be either 'random' or "
                "'append'.")

        # set flag to indicate that the trial was aborted
        self._trialAborted = True  

    def addResponse(self, result, intensity=None):
        """Add a 1 or 0 to signify a correct / detected or
        incorrect / missed trial

        This is essential to advance the staircase to a new intensity level!
        """
        self.currentStaircase.addResponse(result, intensity)
        if self.currentStaircase.finished:
            self.runningStaircases.remove(self.currentStaircase)
        # add the current data to experiment if poss
        if self.getExp() != None:  # update the experiment handler too
            self.getExp().addData(self.name + ".response", result)
        self.totalTrials += 1

    def addOtherData(self, name, value):
        """Add some data about the current trial that will not be used to
        control the staircase(s) such as reaction time data
        """
        self.currentStaircase.addOtherData(name, value)

    def addData(self, result, intensity=None):
        """Deprecated 1.79.00: It was ambiguous whether you were adding
        the response (0 or 1) or some other data concerning the trial so
        there is now a pair of explicit methods:

        *   addResponse(corr,intensity) #some data that alters the next
                trial value
        *   addOtherData('RT', reactionTime) #some other data that won't
                control staircase

        """
        self.addResponse(result, intensity)
        if isinstance(result, str):
            raise TypeError("MultiStairHandler.addData should only receive "
                            "corr / incorr. Use .addOtherData('datName',val)")

    def saveAsPickle(self, fileName, fileCollisionMethod='rename'):
        """Saves a copy of self (with data) to a pickle file.

        This can be reloaded later and further analyses carried out.

        :Parameters:

            fileCollisionMethod: Collision method passed to
            :func:`~psychopy.tools.fileerrortools.handleFileCollision`

        """
        if self.totalTrials < 1:
            if self.autoLog:
                logging.debug('StairHandler.saveAsPickle called but no '
                              'trials completed. Nothing saved')
            return -1

        # otherwise use default location
        if not fileName.endswith('.psydat'):
            fileName += '.psydat'

        with openOutputFile(fileName=fileName, append=False,
                           fileCollisionMethod=fileCollisionMethod) as f:
            pickle.dump(self, f)

        if (fileName is not None) and (fileName != 'stdout'):
            logging.info('saved data to %s' % f.name)

    def saveAsExcel(self, fileName, matrixOnly=False, appendFile=False,
                    fileCollisionMethod='rename'):
        """Save a summary data file in Excel OpenXML format workbook
        (:term:`xlsx`) for processing in most spreadsheet packages.
        This format is compatible with versions of Excel (2007 or greater)
        and with OpenOffice (>=3.0).

        It has the advantage over the simpler text files (see
        :func:`TrialHandler.saveAsText()` )
        that the data from each staircase will be save in the same file, with
        the sheet name coming from the 'label' given in the dictionary of
        conditions during initialisation of the Handler.

        The file extension `.xlsx` will be added if not given already.

        The file will contain a set of values specifying the staircase level
        ('intensity') at each reversal, a list of reversal indices
        (trial numbers), the raw staircase/intensity level on *every* trial
        and the corresponding responses of the participant on every trial.

        :Parameters:

            fileName: string
                the name of the file to create or append. Can include
                relative or absolute path

            matrixOnly: True or False
                If set to True then only the data itself will be output
                (no additional info)

            appendFile: True or False
                If False any existing file with this name will be overwritten.
                If True then a new worksheet will be appended.
                If a worksheet already exists with that name a number will
                be added to make it unique.

            fileCollisionMethod: string
                Collision method passed to
                :func:`~psychopy.tools.fileerrortools.handleFileCollision`
                This is ignored if ``append`` is ``True``.

        """
        if self.totalTrials < 1:
            if self.autoLog:
                logging.debug('StairHandler.saveAsExcel called but no'
                              ' trials completed. Nothing saved')
            return -1

        append = appendFile
        for thisStair in self.staircases:
            # make a filename
            label = thisStair.condition['label']
            thisStair.saveAsExcel(
                fileName, sheetName=label, matrixOnly=matrixOnly,
                appendFile=append, fileCollisionMethod=fileCollisionMethod)
            append = True

    def saveAsText(self, fileName,
                   delim=None,
                   matrixOnly=False,
                   fileCollisionMethod='rename',
                   encoding='utf-8-sig'):
        """Write out text files with the data.

        For MultiStairHandler this will output one file for each staircase
        that was run, with _label added to the fileName that you specify above
        (label comes from the condition dictionary you specified when you
        created the Handler).

        :Parameters:

            fileName: a string
                The name of the file, including path if needed. The extension
                `.tsv` will be added if not included.

            delim: a string
                the delimiter to be used (e.g. '\t' for tab-delimited,
                ',' for csv files)

            matrixOnly: True/False
                If True, prevents the output of the `extraInfo` provided
                at initialisation.

            fileCollisionMethod:
                Collision method passed to
                :func:`~psychopy.tools.fileerrortools.handleFileCollision`

            encoding:
                The encoding to use when saving a the file.
                Defaults to `utf-8-sig`.

        """
        if self.totalTrials < 1:
            if self.autoLog:
                logging.debug('StairHandler.saveAsText called but no trials'
                              ' completed. Nothing saved')
            return -1
        for thisStair in self.staircases:
            # make a filename
            label = thisStair.condition['label']
            thisFileName = fileName + "_" + label
            thisStair.saveAsText(
                fileName=thisFileName, delim=delim, matrixOnly=matrixOnly,
                fileCollisionMethod=fileCollisionMethod, encoding=encoding
            )

    def printAsText(self,
                    delim='\t',
                    matrixOnly=False):
        """Write the data to the standard output stream

        :Parameters:

            delim: a string
                the delimitter to be used (e.g. '\t' for tab-delimitted,
                ',' for csv files)

            matrixOnly: True/False
                If True, prevents the output of the `extraInfo` provided
                at initialisation.
        """
        nStairs = len(self.staircases)
        for stairN, thisStair in enumerate(self.staircases):
            if stairN < nStairs - 1:
                thisMatrixOnly = True  # no header info for first files
            else:
                thisMatrixOnly = matrixOnly
            # make a filename
            label = thisStair.condition['label']
            thisStair.saveAsText(fileName='stdout', delim=delim,
                                 matrixOnly=thisMatrixOnly)


if __name__ == "__main__":
    pass
