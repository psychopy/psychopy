#!/usr/bin/env python
# -*- coding: utf-8 -*-

import sys

from packaging.version import Version

from .base import DataHandler
from .routine import Routine
from .experiment import ExperimentHandler
from .trial import TrialHandler, TrialHandler2, TrialHandlerExt, TrialType
from .staircase import (StairHandler, QuestHandler, PsiHandler,
                        MultiStairHandler)
from .counterbalance import Counterbalancer
from . import shelf

if sys.version_info.major == 3 and sys.version_info.minor >= 6:
    from .staircase import QuestPlusHandler

from .utils import (checkValidFilePath, isValidVariableName, importTrialTypes,
                    sliceFromString, indicesFromString, importConditions,
                    createFactorialTrialList, bootStraps, functionFromStaircase,
                    getDateStr)

from .fit import (FitFunction, FitCumNormal, FitLogistic, FitNakaRushton,
                  FitWeibull)

try:
    # import openpyxl
    import openpyxl
    if Version(openpyxl.__version__) >= Version('2.4.0'):
        # openpyxl moved get_column_letter to utils.cell
        from openpyxl.utils.cell import get_column_letter
    else:
        from openpyxl.cell import get_column_letter
    from openpyxl.reader.excel import load_workbook
    haveOpenpyxl = True
except ImportError:
    haveOpenpyxl = False

try:
    import xlrd
    haveXlrd = True
except ImportError:
    haveXlrd = False
