'''
Data import functions
'''

import cPickle
import re
import numpy
import csv

from matplotlib import mlab

from psychopy.errors import DataImportError, DataFormatError

try:
    import openpyxl.reader.excel
    import openpyxl.cell
    haveOpenpyxl=True
except:
    haveOpenpyxl=False

_nonalphanumeric_re = re.compile(r'\W') # will match all bad var name chars

def isValidVariableName(name):
    """Checks whether a certain string could be used as a valid variable.

    Usage::

        OK, msg = isValidVariableName(name)

    >>> isValidVariableName('name')
    (True, '')
    >>> isValidVariableName('0name')
    (False, 'Variables cannot begin with numeric character')
    >>> isValidVariableName('first second')
    (False, 'Variables cannot contain punctuation or spaces')
    >>> isValidVariableName('')
    (False, "Variables cannot be missing, None, or ''")
    >>> isValidVariableName(None)
    (False, "Variables cannot be missing, None, or ''")
    >>> isValidVariableName(23)
    (False, "Variables must be string-like")
    >>> isValidVariableName('a_b_c')
    (True, '')
    """
    if not name:
        return False, "Variables cannot be missing, None, or ''"
    if not type(name) in [str, unicode, numpy.string_, numpy.unicode_]:
        return False, "Variables must be string-like"
    try:
        name=str(name)#convert from unicode if possible
    except:
        if type(name) in [unicode, numpy.unicode_]:
            raise AttributeError, "name %s (type %s) contains non-ASCII characters (e.g. accents)" % (name, type(name))
        else:
            raise AttributeError, "name %s (type %s) could not be converted to a string" % (name, type(name))

    if name[0].isdigit():
        return False, "Variables cannot begin with numeric character"
    if _nonalphanumeric_re.search(name):
        return False, "Variables cannot contain punctuation or spaces"
    return True, ""

def _assertValidVarNames(fieldNames):
        """screens a list of names as candidate variable names. if all names are
        OK, return silently; else raise ImportError with msg
        """
        if not all(fieldNames):
            raise DataImportError, 'Missing parameter name(s); empty cell(s) in the first row?'
        for name in fieldNames:
            OK, msg = isValidVariableName(name)
            if not OK: #tailor message to importConditions
                msg = msg.replace('Variables', 'Parameters (column headers)')
                raise DataImportError, '%s ("%s")' % (msg, name)

def importPickleFromFile(dataFile):
    try:
        trialsArr = cPickle.load(dataFile)
    except:
        raise DataFormatError, 'Could not load pickle file'
    dataFile.close()
    trialList = []
    fieldNames = trialsArr[0] # header line first
    _assertValidVarNames(fieldNames)
    for row in trialsArr[1:]:
        thisTrial = {}
        for fieldN, fieldName in enumerate(fieldNames):
            thisTrial[fieldName] = row[fieldN] # type is correct, being .pkl
        trialList.append(thisTrial)
    return trialList, fieldNames

def importCSVFromFile(dataFile):
    #use matplotlib to import data and intelligently check for data types
    #all data in one column will be given a single type (e.g. if one cell is string, all will be set to string)
    #convert the record array into a list of dicts
    trialsArr = mlab.csv2rec(dataFile)# data = non-header row x col
    fieldNames = trialsArr.dtype.names
    dataFile.close()
    _assertValidVarNames(fieldNames)
    trialList = []
    for trialN, _ in enumerate(trialsArr):
        thisTrial ={}
        for fieldN, fieldName in enumerate(fieldNames):
            val = trialsArr[trialN][fieldN]
            if type(val)==numpy.string_:
                val = unicode(val.decode('utf-8'))
                #if it looks like a list, convert it:
                if val.startswith('[') and val.endswith(']'):
                    #exec('val=%s' %unicode(val.decode('utf8')))
                    val = eval(val)
            thisTrial[fieldName] = val
        trialList.append(thisTrial)
    return trialList, fieldNames

def _getExcelCellName(col, row):
    """Returns the excel cell name for a row and column (zero-indexed)

    >>> _getExcelCellName(0,0)
    'A1'
    >>> _getExcelCellName(2,1)
    'C2'
    """
    return "%s%i" %(openpyxl.cell.get_column_letter(col+1), row+1)#BEWARE - openpyxl uses indexing at 1, to fit with Excel

def importXLSXFromFile(dataFile):
    if not haveOpenpyxl:
        raise DataFormatError, 'Openpyxl is required for XLSX files, but it was not found.'
    try:
        wb = openpyxl.reader.excel.load_workbook(dataFile)
    except: # InvalidFileException(unicode(e)): # this fails
        raise DataFormatError, 'Could not load XLSX file' 
    ws = wb.worksheets[0]
    nCols = ws.get_highest_column()
    nRows = ws.get_highest_row()

    #get parameter names from the first row header
    fieldNames = []
    for colN in range(nCols):
        fieldName = ws.cell(_getExcelCellName(col=colN, row=0)).value
        fieldNames.append(fieldName)
    _assertValidVarNames(fieldNames)

    #loop trialTypes
    trialList = []
    for rowN in range(1, nRows):#skip header first row
        thisTrial={}
        for colN in range(nCols):
            val = ws.cell(_getExcelCellName(col=colN, row=rowN)).value
            #if it looks like a list, convert it
            if type(val) in [unicode, str] and (
                    val.startswith('[') and val.endswith(']') or
                    val.startswith('(') and val.endswith(')') ):
                val = eval(val)
            fieldName = fieldNames[colN]
            thisTrial[fieldName] = val
        trialList.append(thisTrial)

FORMATS = {
    ".csv": importCSVFromFile,
    ".pkl": importPickleFromFile,
    ".xlsx": importXLSXFromFile 
}
