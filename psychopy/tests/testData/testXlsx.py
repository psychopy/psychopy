"""Tests for psychopy.data.DataHandler"""
import os, shutil, tempfile
import nose
import numpy

from openpyxl.reader.excel import load_workbook
from psychopy import data, misc

thisDir,filename = os.path.split(os.path.abspath(__file__))

class TestXLSX:
    def setUp(self):
        self.name = 'testXlsx'
        self.fullName = self.name+'.xlsx'
        
    def tearDown(self):
        os.remove(self.fullName)
        
    def testTrialHandlerAndXLSX(self):
        conds = data.importConditions(os.path.join(thisDir, 'trialTypes.xlsx'))
        trials = data.TrialHandler(trialList=conds, seed=100, nReps=2)
        responses=[1,1,2,3,2,3, 1,3,2,2,1,1]
        rts=numpy.array(responses)/10.0
        for trialN, trial in enumerate(trials):
            trials.addData('resp', responses[trialN])
            trials.addData('rt',rts[trialN])
        trials.saveAsExcel(self.name)
        
        # Make sure the file is there
        assert os.path.isfile(self.fullName)
        expBook = load_workbook(os.path.join(thisDir,'corrXlsx.xlsx'))
        actBook = load_workbook(self.fullName)
        
        for wsN, expWS in enumerate(expBook.worksheets):
            actWS = actBook.worksheets[wsN]
            for key, expVal in expWS._cells.items():
                actVal = actWS._cells[key]
                try:
                    # convert to float if possible and compare with a reasonable
                    # (default) precision
                    expVal.value = float(expVal.value)
                    nose.tools.assert_almost_equals(expVal.value,
                                                    float(actVal.value))
                except:
                    # otherwise do precise comparison
                    nose.tools.assert_equal(expVal.value, actVal.value)

def testTrialTypeImport():
    fromCSV = data.importConditions(os.path.join(thisDir, 'trialTypes.csv'))
    fromXLSX = data.importConditions(os.path.join(thisDir, 'trialTypes.xlsx'))
    
    for trialN, trialCSV in enumerate(fromCSV):
        trialXLSX = fromXLSX[trialN]
        assert trialXLSX.keys()==trialCSV.keys()
        for header in trialCSV.keys():
            if trialXLSX[header]==None and numpy.isnan(trialCSV[header]):
                trialCSV[header]=None#this is ok
            if trialXLSX[header] != trialCSV[header]:
                print header, trialCSV[header], trialXLSX[header]
            assert trialXLSX[header] == trialCSV[header]
