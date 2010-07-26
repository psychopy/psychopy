"""Tests for psychopy.data.DataHandler"""
import os
import shutil
import tempfile

from openpyxl.reader.excel import load_workbook
from psychopy import data, misc

thisDir,filename = os.path.split(os.path.abspath(__file__))
name = 'psychopy_testXlsxOutput'
fullName = name+'.xlsx'
class TestTrialHandler:
    def setUp(self):
        pass
        
    def tearDown(self):
        os.remove(fullName)

    def test(self):
        dat = misc.fromFile(os.path.join(thisDir, 'data.psydat'))
        dat.saveAsExcel(name,
            stimOut=['text', 'congruent', 'corrAns', 'letterColor', ],
            dataOut=['n','all_mean','all_std', 'all_raw'])
        
        # Make sure the file is there
        assert os.path.isfile(fullName)
        
        expBook = load_workbook(os.path.join(thisDir,'data.xlsx'))
        actBook = load_workbook(fullName)
        
        for wsN, expWS in enumerate(expBook.worksheets):
            actWS = actBook.worksheets[wsN]
            for key, expVal in expWS._cells.items():
                actVal = actWS._cells[key]
                print actVal.value, expVal.value
                assert actVal.value == expVal.value
