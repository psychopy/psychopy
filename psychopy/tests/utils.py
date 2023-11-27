import sys
from os.path import abspath, basename, dirname, isfile, isdir, join as pjoin
import os.path
from pathlib import Path
import shutil
import numpy as np
import io
from psychopy import logging, colors

try:
    from PIL import Image
except ImportError:
    import Image

import pytest

# define the path where to find testing data
# so tests could be ran from any location
TESTS_PATH = abspath(dirname(__file__))
TESTS_DATA_PATH = pjoin(TESTS_PATH, 'data')
TESTS_FAILS_PATH = pjoin(TESTS_PATH, 'fails', sys.platform)
TESTS_FONT = pjoin(TESTS_DATA_PATH, 'DejaVuSerif.ttf')

# Make sure all paths exist
if not isdir(TESTS_FAILS_PATH):
    os.makedirs(TESTS_FAILS_PATH)

# Some regex shorthand
_q = r"[\"']"  # quotes
_lb = r"[\[\(]"  # left bracket
_rb = r"[\]\)]"  # right bracket
_d = r"\$"  # dollar (escaped for re)
_sl = r"\\\\"  # back slash


def getFailFilenames(fileName, tag=""):
    """
    Create variant of given filename for a failed test

    Parameters
    ==========
    fileName : str or Path
        Path to original file
    tag : str
        Optional tag to append to the file stem

    Returns
    ==========
    str
        Path to the local copy
    str
        Path to the copy of the exemplar in the fails folder
    """
    # Path-ise filename
    fileName = Path(fileName)
    # Create new stem
    if tag:
        tag = "_" + tag
    stem = fileName.stem + tag
    # Construct new filename for local copy
    localFileName = pjoin(TESTS_FAILS_PATH, stem + "_local" + fileName.suffix)
    # Construct new filename for exemplar copy
    exemplarFileName = pjoin(TESTS_FAILS_PATH, fileName.stem + fileName.suffix)

    return localFileName, exemplarFileName


def compareScreenshot(fileName, win, tag="", crit=5.0):
    """Compare the current back buffer of the given window with the file

    Screenshots are stored and compared against the files under path
    kept in TESTS_DATA_PATH.  Thus specify relative path to that
    directory
    """
    #if we start this from a folder below run.py the data folder won't be found
    fileName = pjoin(TESTS_DATA_PATH, fileName)
    #get the frame from the window
    win.getMovieFrame(buffer='back')
    frame=win.movieFrames[-1]
    win.movieFrames=[]
    #if the file exists run a test, if not save the file
    if not isfile(fileName):
        frame = frame.resize((int(frame.size[0]/2), int(frame.size[1]/2)),
                             resample=Image.LANCZOS)
        frame.save(fileName, optimize=1)
        pytest.skip("Created %s" % basename(fileName))
    else:
        expected = Image.open(fileName)
        expDat = np.array(expected.getdata())
        imgDat = np.array(frame.getdata())
        # for retina displays the frame data is 4x bigger than expected
        if imgDat.shape[0] != expDat.shape[0]:
            frame = frame.resize(expected.size, resample=Image.LANCZOS)
            imgDat = np.array(frame.getdata())
            crit += 5  # be more relaxed because of the interpolation
        rms = np.std(imgDat-expDat)
        localFileName, exemplarFileName = getFailFilenames(fileName, tag=tag)
        if rms >= crit/2:
            #there was SOME discrepancy
            logging.warning('PsychoPyTests: RMS=%.3g at threshold=%3.g'
                  % (rms, crit))
        if not rms<crit: #don't do `if rms>=crit because that doesn't catch rms=nan
            # If test fails, save local copy and copy of exemplar to fails folder
            frame.save(localFileName, optimize=1)
            expected.save(exemplarFileName, optimize=1)
            logging.warning('PsychoPyTests: Saving local copy into %s' % localFileName)
        assert rms<crit, \
            "RMS=%.3g at threshold=%.3g. Local copy in %s" % (rms, crit, localFileName)


def compareTextFiles(pathToActual, pathToCorrect, delim=None,
                     encoding='utf-8-sig', tolerance=None):
    """Compare the text of two files, ignoring EOL differences,
    and save a copy if they differ

    State a tolerance, or percentage of errors allowed,
    to account for differences in version numbers, datetime, etc
    """

    if not os.path.isfile(pathToCorrect):
        logging.warning('There was no comparison ("correct") file available, for path "{pathToActual}"\n'
                        '\t\t\tSaving current file as the comparison: {pathToCorrect}'
                        .format(pathToActual=pathToActual,
                                pathToCorrect=pathToCorrect))
        shutil.copyfile(pathToActual, pathToCorrect)
        raise IOError("File not found")  # deliberately raise an error to see the warning message, but also to create file

    allowLines = 0
    nLinesMatch = True

    if delim is None:
        if pathToCorrect.endswith('.csv'):
            delim=','
        elif pathToCorrect.endswith(('.dlm', '.tsv')):
            delim='\t'

    try:
        # we have the necessary file
        with io.open(pathToActual, 'r', encoding='utf-8-sig', newline=None) as f:
            txtActual = f.readlines()

        with io.open(pathToCorrect, 'r', encoding='utf-8-sig', newline=None) as f:
            txtCorrect = f.readlines()

        if tolerance is not None:
            # Set number of lines allowed to fail
            allowLines = round((tolerance * len(txtCorrect)) / 100, 0)

        # Check number of lines per document for equality
        nLinesMatch = len(txtActual) == len(txtCorrect)
        assert nLinesMatch
        errLines = []

        for lineN in range(len(txtActual)):
            if delim is None:
                lineActual = txtActual[lineN]
                lineCorrect = txtCorrect[lineN]

                # just compare the entire line
                if not lineActual == lineCorrect:
                    errLines.append({'actual':lineActual, 'correct':lineCorrect})
                assert len(errLines) <= allowLines

            else:  # word by word instead
                lineActual=txtActual[lineN].split(delim)
                lineCorrect=txtCorrect[lineN].split(delim)

                for wordN in range(len(lineActual)):
                    wordActual=lineActual[wordN]
                    wordCorrect=lineCorrect[wordN]
                    try:
                        wordActual=float(wordActual.lstrip('"[').strip(']"'))
                        wordCorrect=float(wordCorrect.lstrip('"[').strip(']"'))
                        # its not a whole well-formed list because .split(delim)
                        isFloat=True
                    except Exception:#stick with simple text if not a float value
                        isFloat=False
                        pass
                    if isFloat:
                        #to a default of 8 dp?
                        assert np.allclose(wordActual,wordCorrect), "Numeric values at (%i,%i) differ: %f != %f " \
                            %(lineN, wordN, wordActual, wordCorrect)
                    else:
                        if wordActual!=wordCorrect:
                            print('actual:')
                            print(repr(txtActual[lineN]))
                            print(lineActual)
                            print('expected:')
                            print(repr(txtCorrect[lineN]))
                            print(lineCorrect)
                        assert wordActual==wordCorrect, "Values at (%i,%i) differ: %s != %s " \
                            %(lineN, wordN, repr(wordActual), repr(wordCorrect))

    except AssertionError as err:
        pathToLocal, pathToExemplar = getFailFilenames(pathToCorrect)

        # Set assertion type
        if not nLinesMatch:  # Fail if number of lines not equal
            msg = "{} has the wrong number of lines".format(pathToActual)
        elif len(errLines) < allowLines:  # Fail if tolerance reached
            msg = 'Number of differences in {failed} exceeds the {tol}% tolerance'.format(failed=pathToActual,
                                                                                          tol=tolerance or 0)
        else:
            shutil.copyfile(pathToActual, pathToLocal)
            shutil.copyfile(pathToCorrect, pathToExemplar)
            msg = "txtActual != txtCorr: Saving local copy to {}".format(pathToLocal)
        logging.error(msg)
        raise AssertionError(err)


def compareXlsxFiles(pathToActual, pathToCorrect):
    from openpyxl.reader.excel import load_workbook
    # Make sure the file is there
    expBook = load_workbook(pathToCorrect)
    actBook = load_workbook(pathToActual)
    error=None

    for wsN, expWS in enumerate(expBook.worksheets):
        actWS = actBook.worksheets[wsN]
        for key, expVal in list(expWS._cells.items()):
            actVal = actWS._cells[key].value
            expVal = expVal.value
            # intercept lists-of-floats, which might mismatch by rounding error
            isListableFloatable = False
            if u"{}".format(expVal).startswith('['):
                expValList = eval(u"{}".format(expVal))
                try:
                    expVal = np.array(expValList, dtype=float)
                    actVal = np.array(eval(u"{}".format(actVal)),
                                      dtype=float) # should go through if expVal does...
                    isListableFloatable = True
                except Exception:
                    pass # non-list+float-able at this point = default
            #determine whether there will be errors
            try:
                # convert to float if possible and compare with a reasonable
                # (default) precision
                expVal = float(expVal)
                isFloatable=True
            except Exception:
                isFloatable=False
            if isListableFloatable:
                if not np.allclose(expVal, actVal):
                    error = "l+f Cell %s: %f != %f" %(key, expVal, actVal)
                    break
            elif isFloatable and abs(expVal-float(actVal))>0.0001:
                error = "f Cell %s: %f != %f" %(key, expVal, actVal)
                break
            elif not isFloatable and expVal!=actVal:
                error = "nf Cell %s: %s != %s" %(key, expVal, actVal)
                break
    if error:
        pathToLocal, pathToExemplar = getFailFilenames(pathToCorrect)
        shutil.copyfile(pathToActual, pathToLocal)
        shutil.copyfile(pathToCorrect, pathToExemplar)
        logging.warning("xlsxActual!=xlsxCorr: Saving local copy to %s" % pathToLocal)
        raise IOError(error)


def comparePixelColor(screen, color, coord=(0, 0), context="color_comparison"):
    ogCoord = coord
    # Adjust for retina
    coord = tuple(int(c * screen.getContentScaleFactor()) for c in ogCoord)

    if hasattr(screen, 'getMovieFrame'):  # check it is a Window class (without importing visual in this file)
        # If given a window, get frame from window
        screen.getMovieFrame(buffer='back')
        frame = screen.movieFrames[-1]
        screen.movieFrames = []

    elif isinstance(screen, str):
        # If given a filename, get frame from file
        frame = Image.open(screen)
    else:
        # If given anything else, throw error
        raise TypeError("Function comparePixelColor expected first input of type psychopy.visual.Window or str, received %s" % (type(screen)))
    frameArr = np.array(frame)
    # If given a Color object, convert to rgb255 (this is what PIL uses)
    if isinstance(color, colors.Color):
        color = color.rgb255
    color = np.array(color)
    pixCol = frameArr[coord]
    # Compare observed color to desired color
    closeEnough = True
    for i in range(min(pixCol.size, color.size)):
        closeEnough = closeEnough and abs(pixCol[i] - color[i]) <= 1 # Allow for 1/255 lenience due to rounding up/down in rgb255
    # Assert
    cond = all(c for c in color == pixCol) or closeEnough
    if not cond:
        frame.save(Path(TESTS_FAILS_PATH) / (context + ".png"))
        raise AssertionError(f"Pixel color {pixCol} at {ogCoord} (x{screen.getContentScaleFactor()}) not equal to target color {color}")


def forceBool(value, handler=any):
    """
    Force a value to a boolean, accounting for the possibility that it is an
    array of booleans.

    Parameters
    ----------
    value
        Value to force
    mode : str
        Method to apply to values which fail builtin `bool` function, e.g. `any`
        or `all` for boolean arrays.

    Returns
    -------
    bool
    """
    try:
        # attempt to make bool
        value = bool(value)
    except ValueError:
        # if this fails,
        value = handler(value)

    return value
